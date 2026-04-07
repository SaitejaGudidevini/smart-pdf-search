"""CellDNA — extract the hidden truth from Excel cells via openpyxl.

Every .xlsx file is a ZIP of XML. openpyxl exposes the full cell metadata:
number formats, fonts, fills, merges, comments, data validation, and formulas.

CellDNA captures ALL of this per-cell so the enricher and LLM get semantic
context instead of raw values. When the number format says '$#,##0.00', we
KNOW it's USD currency — no LLM guess needed.

Usage:
    inspector = WorkbookDNA(file_path)
    for sheet_name, sheet_dna in inspector.sheets.items():
        for col_name, col_dna in sheet_dna.column_dna.items():
            print(col_dna.semantic_type)  # 'currency_usd', 'percentage', 'date', ...
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


# ======================================================================
# Built-in Excel number format IDs
# ======================================================================

# Excel stores common formats as integer IDs (not strings).
# Full list: https://learn.microsoft.com/en-us/dotnet/api/documentformat.openxml.spreadsheet.numberingformat
_BUILTIN_FORMATS: dict[int, str] = {
    0: "general",
    1: "integer",       # 0
    2: "decimal",       # 0.00
    3: "integer",       # #,##0
    4: "decimal",       # #,##0.00
    5: "currency",      # $#,##0
    6: "currency",      # $#,##0 (red negative)
    7: "currency",      # $#,##0.00
    8: "currency",      # $#,##0.00 (red negative)
    9: "percentage",    # 0%
    10: "percentage",   # 0.00%
    11: "scientific",   # 0.00E+00
    12: "fraction",     # # ?/?
    13: "fraction",     # # ??/??
    14: "date",         # m/d/yyyy
    15: "date",         # d-mmm-yy
    16: "date",         # d-mmm
    17: "date",         # mmm-yy
    18: "time",         # h:mm AM/PM
    19: "time",         # h:mm:ss AM/PM
    20: "time",         # h:mm
    21: "time",         # h:mm:ss
    22: "datetime",     # m/d/yyyy h:mm
    37: "accounting",   # #,##0;(#,##0)
    38: "accounting",   # #,##0;[Red](#,##0)
    39: "accounting",   # #,##0.00;(#,##0.00)
    40: "accounting",   # #,##0.00;[Red](#,##0.00)
    41: "accounting",   # _(* #,##0_)
    42: "currency",     # _($* #,##0_)
    43: "accounting",   # _(* #,##0.00_)
    44: "currency",     # _($* #,##0.00_)
    45: "time",         # mm:ss
    46: "time",         # [h]:mm:ss
    47: "time",         # mm:ss.0
    48: "scientific",   # ##0.0E+0
    49: "text",         # @
}

# Currency symbols we recognize in format strings
_CURRENCY_SYMBOLS = {
    "$": "currency_usd",
    "€": "currency_eur",
    "£": "currency_gbp",
    "¥": "currency_jpy",
    "₹": "currency_inr",
    "₩": "currency_krw",
    "₣": "currency_chf",
    "R$": "currency_brl",
    "kr": "currency_sek",
}


# ======================================================================
# Data classes
# ======================================================================

@dataclass
class CellDNA:
    """Everything openpyxl knows about a single cell."""
    value: Any
    row: int
    col: int
    coordinate: str           # e.g. "B5"

    # Format intelligence
    number_format: str        # raw format string: '$#,##0.00' or 'General'
    semantic_type: str        # interpreted: 'currency_usd', 'percentage', 'date', etc.

    # Style signals
    is_bold: bool = False
    is_italic: bool = False
    font_size: float | None = None
    font_color: str | None = None     # hex color (6 chars)
    fill_color: str | None = None     # background hex color
    indent_level: int = 0             # 0 = top-level, 1+ = sub-item

    # Structure signals
    is_merged: bool = False
    merge_range: str | None = None    # e.g. "A1:D1"
    formula: str | None = None        # raw formula string
    comment: str | None = None        # author notes

    # Data validation (dropdown rules, allowed values)
    validation_type: str | None = None   # "list", "whole", "decimal", etc.
    validation_formula: str | None = None  # e.g. '"Yes,No"' or 'Sheet2!$A$1:$A$5'

    # Computed flags
    is_header_candidate: bool = False  # bold + text + near top
    is_total_candidate: bool = False   # bold + SUM/TOTAL formula
    is_highlighted: bool = False       # non-white background


@dataclass
class ColumnDNA:
    """Aggregated DNA for an entire column — the column's "identity"."""
    column_name: str
    column_index: int         # 0-based

    # The dominant semantic type across all data cells in this column
    semantic_type: str        # 'currency_usd', 'percentage', 'date', 'text', etc.
    number_format: str        # most common non-General format in the column

    # Inferred role (from format + style patterns, NOT from LLM)
    inferred_role: str | None = None   # 'metric', 'temporal', 'identifier', etc.
    format_hint: str = ""              # passthrough to ColumnClassification

    # Style patterns across the column
    has_bold_header: bool = False
    has_total_row: bool = False
    has_formulas: bool = False
    has_validation: bool = False
    indent_levels: set = field(default_factory=set)  # hierarchy detection

    # Data validation values (if column has a dropdown list)
    validation_values: list[str] = field(default_factory=list)

    # Confidence: how many cells agreed on the semantic type
    type_confidence: float = 0.0  # 0.0-1.0

    def to_dict(self) -> dict:
        return {
            "column_name": self.column_name,
            "semantic_type": self.semantic_type,
            "number_format": self.number_format,
            "inferred_role": self.inferred_role,
            "format_hint": self.format_hint,
            "has_bold_header": self.has_bold_header,
            "has_total_row": self.has_total_row,
            "has_formulas": self.has_formulas,
            "has_validation": self.has_validation,
            "validation_values": self.validation_values,
            "type_confidence": round(self.type_confidence, 2),
        }


@dataclass
class SheetDNA:
    """Aggregated DNA for an entire sheet."""
    sheet_name: str
    column_dna: dict[str, ColumnDNA] = field(default_factory=dict)

    # Sheet-level signals
    merged_ranges: list[str] = field(default_factory=list)
    named_tables: list[dict] = field(default_factory=list)  # Excel Table objects
    has_auto_filter: bool = False
    has_conditional_formatting: bool = False
    frozen_panes: str | None = None  # e.g. "A2" = header row frozen

    # Detected structure
    header_rows: list[int] = field(default_factory=list)   # rows that are bold
    total_rows: list[int] = field(default_factory=list)     # rows with SUM formulas
    indent_hierarchy: bool = False  # True if indent-based tree structure detected

    def to_dict(self) -> dict:
        return {
            "sheet_name": self.sheet_name,
            "columns": {k: v.to_dict() for k, v in self.column_dna.items()},
            "merged_ranges": self.merged_ranges,
            "named_tables": self.named_tables,
            "has_auto_filter": self.has_auto_filter,
            "has_conditional_formatting": self.has_conditional_formatting,
            "frozen_panes": self.frozen_panes,
            "header_rows": self.header_rows,
            "total_rows": self.total_rows,
            "indent_hierarchy": self.indent_hierarchy,
        }


# ======================================================================
# Core extractor
# ======================================================================

class WorkbookDNA:
    """Extract the full DNA of an Excel workbook.

    Opens the file twice via openpyxl:
      - data_only=True  → computed cell values
      - data_only=False → raw formulas

    Then walks every cell to build CellDNA, aggregates into ColumnDNA
    and SheetDNA.
    """

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.sheets: dict[str, SheetDNA] = {}

        self._wb_values = openpyxl.load_workbook(file_path, data_only=True)
        self._wb_formulas = None
        try:
            self._wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        except Exception:
            pass

        self._extract_all()

    def close(self):
        self._wb_values.close()
        if self._wb_formulas:
            self._wb_formulas.close()

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()

    # ------------------------------------------------------------------
    # Main extraction
    # ------------------------------------------------------------------

    def _extract_all(self):
        for sheet_name in self._wb_values.sheetnames:
            ws = self._wb_values[sheet_name]
            if ws.sheet_state != "visible":
                continue

            ws_formulas = None
            if self._wb_formulas and sheet_name in self._wb_formulas.sheetnames:
                ws_formulas = self._wb_formulas[sheet_name]

            sheet_dna = self._extract_sheet(ws, ws_formulas, sheet_name)
            self.sheets[sheet_name] = sheet_dna

    def _extract_sheet(
        self, ws: Worksheet, ws_formulas: Worksheet | None, sheet_name: str,
    ) -> SheetDNA:
        """Extract DNA for a single sheet."""
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            return SheetDNA(sheet_name=sheet_name)

        # Pre-compute merged cell lookup
        merged_lookup: dict[str, str] = {}
        merged_ranges: list[str] = []
        for mr in ws.merged_cells.ranges:
            range_str = str(mr)
            merged_ranges.append(range_str)
            for cell in mr.cells:
                # cell is (row, col) tuple
                coord = f"{_col_letter(cell[1])}{cell[0]}"
                merged_lookup[coord] = range_str

        # Pre-compute data validation lookup
        validation_lookup: dict[str, tuple[str, str]] = {}
        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                vtype = dv.type or "unknown"
                vformula = str(dv.formula1) if dv.formula1 else ""
                if dv.sqref:
                    for cell_range in str(dv.sqref).split():
                        # Could be a range like "B2:B100" or a single cell "B2"
                        if ":" in cell_range:
                            # Expand range — just store the info, we'll match by column
                            validation_lookup[cell_range] = (vtype, vformula)
                        else:
                            validation_lookup[cell_range] = (vtype, vformula)

        # Collect all CellDNA per column
        # column_cells[col_index] = list of CellDNA for data rows
        column_cells: dict[int, list[CellDNA]] = {}
        header_row_dna: dict[int, CellDNA] = {}  # col_index -> header cell DNA
        all_row_bold: dict[int, bool] = {}  # row_index -> all cells bold?
        all_row_has_sum: dict[int, bool] = {}

        # Detect header row (first row with mostly bold text cells)
        header_row_idx = self._detect_header_row(ws, max_row, max_col)

        for row in ws.iter_rows(min_row=1, max_row=max_row,
                                min_col=1, max_col=max_col):
            row_idx = row[0].row
            row_bold_count = 0
            row_cell_count = 0
            row_has_sum = False

            for cell in row:
                if cell.value is None and not (cell.has_style and cell.font and cell.font.bold):
                    continue

                dna = self._extract_cell(cell, ws_formulas, merged_lookup, validation_lookup)

                col_idx = cell.column - 1
                row_cell_count += 1
                if dna.is_bold:
                    row_bold_count += 1
                if dna.formula and re.search(r"\bSUM\b", dna.formula, re.IGNORECASE):
                    row_has_sum = True

                if row_idx == header_row_idx:
                    header_row_dna[col_idx] = dna
                elif row_idx > header_row_idx:
                    column_cells.setdefault(col_idx, []).append(dna)

            if row_cell_count > 0:
                all_row_bold[row_idx] = row_bold_count == row_cell_count
                all_row_has_sum[row_idx] = row_has_sum

        # Detect header and total rows
        header_rows = [r for r, is_bold in all_row_bold.items()
                       if is_bold and r <= header_row_idx]
        total_rows = [r for r, has_sum in all_row_has_sum.items() if has_sum]

        # Build ColumnDNA from aggregated CellDNA
        column_dna: dict[str, ColumnDNA] = {}
        for col_idx, cells in column_cells.items():
            header = header_row_dna.get(col_idx)
            col_name = str(header.value).strip() if header and header.value else f"col_{col_idx}"
            cdna = self._aggregate_column(col_name, col_idx, cells, header)
            column_dna[col_name] = cdna

        # Sheet-level signals
        named_tables = []
        for table in ws.tables.values():
            named_tables.append({
                "name": table.displayName,
                "ref": table.ref,
            })

        has_filter = ws.auto_filter is not None and ws.auto_filter.ref is not None
        has_cond_fmt = bool(ws.conditional_formatting)
        frozen = str(ws.freeze_panes) if ws.freeze_panes else None

        # Detect indent hierarchy
        all_indents: set[int] = set()
        for cells in column_cells.values():
            for c in cells:
                if c.indent_level > 0:
                    all_indents.add(c.indent_level)

        return SheetDNA(
            sheet_name=sheet_name,
            column_dna=column_dna,
            merged_ranges=merged_ranges,
            named_tables=named_tables,
            has_auto_filter=has_filter,
            has_conditional_formatting=has_cond_fmt,
            frozen_panes=frozen,
            header_rows=header_rows,
            total_rows=total_rows,
            indent_hierarchy=len(all_indents) >= 2,
        )

    # ------------------------------------------------------------------
    # Cell-level extraction
    # ------------------------------------------------------------------

    def _extract_cell(
        self,
        cell: Cell,
        ws_formulas: Worksheet | None,
        merged_lookup: dict[str, str],
        validation_lookup: dict[str, tuple[str, str]],
    ) -> CellDNA:
        """Extract the full DNA of a single cell."""

        # --- Number format → semantic type ---
        nf = cell.number_format or "General"
        semantic_type = infer_semantic_type(nf, cell.data_type, cell.value)

        # --- Style signals ---
        font = cell.font
        fill = cell.fill

        is_bold = bool(font and font.bold)
        is_italic = bool(font and font.italic)
        font_size = font.size if font and font.size else None
        font_color = _color_to_hex(font.color) if font and font.color else None
        fill_color = _color_to_hex(fill.fgColor) if fill and fill.fgColor else None

        indent_level = 0
        if cell.alignment and cell.alignment.indent:
            indent_level = int(cell.alignment.indent)

        # --- Merged cell ---
        coord = cell.coordinate
        is_merged = coord in merged_lookup
        merge_range = merged_lookup.get(coord)

        # --- Formula (from the non-data_only workbook) ---
        formula = None
        if ws_formulas:
            try:
                formula_cell = ws_formulas[coord]
                if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                    formula = formula_cell.value
            except Exception:
                pass

        # --- Comment ---
        comment = cell.comment.text.strip() if cell.comment else None

        # --- Data validation ---
        val_type, val_formula = None, None
        if coord in validation_lookup:
            val_type, val_formula = validation_lookup[coord]
        else:
            # Check range-based validations
            for vrange, (vt, vf) in validation_lookup.items():
                if ":" in vrange and _coord_in_range(coord, vrange):
                    val_type, val_formula = vt, vf
                    break

        # --- Computed flags ---
        is_header = is_bold and cell.row <= 3 and isinstance(cell.value, str)
        is_total = is_bold and formula is not None and bool(
            re.search(r"\b(SUM|TOTAL|SUBTOTAL)\b", formula or "", re.IGNORECASE)
        )
        is_highlighted = (
            fill_color is not None
            and fill_color not in ("000000", "FFFFFF", "")
        )

        return CellDNA(
            value=cell.value,
            row=cell.row,
            col=cell.column,
            coordinate=coord,
            number_format=nf,
            semantic_type=semantic_type,
            is_bold=is_bold,
            is_italic=is_italic,
            font_size=font_size,
            font_color=font_color,
            fill_color=fill_color,
            indent_level=indent_level,
            is_merged=is_merged,
            merge_range=merge_range,
            formula=formula,
            comment=comment,
            validation_type=val_type,
            validation_formula=val_formula,
            is_header_candidate=is_header,
            is_total_candidate=is_total,
            is_highlighted=is_highlighted,
        )

    # ------------------------------------------------------------------
    # Column-level aggregation
    # ------------------------------------------------------------------

    def _aggregate_column(
        self,
        col_name: str,
        col_index: int,
        cells: list[CellDNA],
        header: CellDNA | None,
    ) -> ColumnDNA:
        """Aggregate CellDNA list into a single ColumnDNA."""
        if not cells:
            return ColumnDNA(
                column_name=col_name, column_index=col_index,
                semantic_type="unknown", number_format="General",
            )

        # Count semantic types (excluding 'general' and 'unknown')
        type_counts: dict[str, int] = {}
        format_counts: dict[str, int] = {}
        for c in cells:
            if c.value is not None:
                type_counts[c.semantic_type] = type_counts.get(c.semantic_type, 0) + 1
                if c.number_format != "General":
                    format_counts[c.number_format] = format_counts.get(c.number_format, 0) + 1

        # Dominant type (most common non-general type)
        non_general = {k: v for k, v in type_counts.items() if k not in ("general", "unknown")}
        if non_general:
            dominant_type = max(non_general, key=non_general.get)
            total_typed = sum(type_counts.values())
            confidence = non_general[dominant_type] / total_typed if total_typed > 0 else 0
        else:
            dominant_type = max(type_counts, key=type_counts.get) if type_counts else "unknown"
            confidence = 1.0 if len(type_counts) == 1 else 0.5

        # Most common non-General format
        dominant_format = max(format_counts, key=format_counts.get) if format_counts else "General"

        # Infer role from semantic type
        inferred_role = _type_to_role(dominant_type)

        # Format hint for the enricher
        format_hint = dominant_type

        # Flags
        has_bold_header = bool(header and header.is_bold)
        has_total = any(c.is_total_candidate for c in cells)
        has_formulas = any(c.formula is not None for c in cells)
        has_validation = any(c.validation_type is not None for c in cells)

        # Extract validation list values
        validation_values: list[str] = []
        if has_validation:
            for c in cells:
                if c.validation_type == "list" and c.validation_formula:
                    # Parse comma-separated list: '"Yes,No,Maybe"'
                    raw = c.validation_formula.strip('"')
                    validation_values = [v.strip() for v in raw.split(",") if v.strip()]
                    break

        # Indent levels
        indent_levels = {c.indent_level for c in cells if c.indent_level > 0}

        return ColumnDNA(
            column_name=col_name,
            column_index=col_index,
            semantic_type=dominant_type,
            number_format=dominant_format,
            inferred_role=inferred_role,
            format_hint=format_hint,
            has_bold_header=has_bold_header,
            has_total_row=has_total,
            has_formulas=has_formulas,
            has_validation=has_validation,
            indent_levels=indent_levels,
            validation_values=validation_values,
            type_confidence=confidence,
        )

    # ------------------------------------------------------------------
    # Header row detection
    # ------------------------------------------------------------------

    def _detect_header_row(self, ws: Worksheet, max_row: int, max_col: int) -> int:
        """Detect the header row index (1-based) using style + content heuristics.

        Looks for the first row where:
          - Most cells are bold, OR
          - Most cells are text strings with the row below being numeric
        """
        check_rows = min(5, max_row)
        best_row = 1
        best_score = -1

        for r in range(1, check_rows + 1):
            bold_count = 0
            text_count = 0
            total = 0

            for c in range(1, min(max_col + 1, 50)):
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    continue
                total += 1
                if cell.font and cell.font.bold:
                    bold_count += 1
                if isinstance(cell.value, str):
                    text_count += 1

            if total == 0:
                continue

            # Score: bold cells + text cells (headers are typically bold strings)
            score = (bold_count / total) * 0.6 + (text_count / total) * 0.4
            if score > best_score and total >= 2:
                best_score = score
                best_row = r

        return best_row


# ======================================================================
# Public helpers
# ======================================================================

def infer_semantic_type(
    number_format: str, data_type: str | None = None, value: Any = None,
) -> str:
    """Convert an Excel number format string to a semantic type.

    This is the core "hidden truth" decoder. Instead of the LLM guessing
    whether a column is currency or percentage, we read it directly from
    the format the spreadsheet author chose.

    Returns one of:
        currency_usd, currency_eur, currency_gbp, currency_jpy, currency_inr,
        currency (generic), percentage, date, time, datetime,
        accounting, integer, decimal, scientific, fraction,
        text, number, general, unknown
    """
    if not number_format or number_format == "General":
        # Fall back to data_type / value type
        if data_type == "s" or isinstance(value, str):
            return "text"
        if data_type == "n" or isinstance(value, (int, float)):
            return "number"
        if data_type == "d":
            return "date"
        if data_type == "b":
            return "boolean"
        return "general"

    nf = number_format

    # Check builtin IDs first (openpyxl sometimes stores these as strings)
    try:
        nf_id = int(nf)
        if nf_id in _BUILTIN_FORMATS:
            result = _BUILTIN_FORMATS[nf_id]
            # Refine currency to specific symbol if value context available
            if result == "currency":
                return "currency_usd"  # builtin IDs 5-8, 42, 44 are USD
            return result
    except (ValueError, TypeError):
        pass

    nf_lower = nf.lower()

    # Currency detection — check for symbol in format string
    for symbol, ctype in _CURRENCY_SYMBOLS.items():
        if symbol in nf:
            return ctype

    # Accounting format (underscore + parentheses pattern)
    if re.search(r"_\(\s*[\$€£¥]", nf) or "_)" in nf:
        # Try to detect currency from the format
        for symbol, ctype in _CURRENCY_SYMBOLS.items():
            if symbol in nf:
                return ctype
        return "accounting"

    # Percentage
    if "%" in nf:
        return "percentage"

    # Date patterns
    if re.search(r"[ymd]{2,}", nf_lower) and not re.search(r"[hs]", nf_lower):
        return "date"

    # Time patterns
    if re.search(r"[hs]", nf_lower) and re.search(r":", nf_lower):
        if re.search(r"[ymd]", nf_lower):
            return "datetime"
        return "time"

    # Scientific notation
    if re.search(r"e\+?0", nf_lower):
        return "scientific"

    # Fraction
    if re.search(r"\?/\?", nf):
        return "fraction"

    # Text format
    if nf == "@":
        return "text"

    # Generic number with decimals
    if re.search(r"0\.0", nf):
        return "decimal"

    # Generic integer format
    if re.search(r"#,?##0(?!\.)", nf):
        return "integer"

    # Fallback
    if data_type == "s" or isinstance(value, str):
        return "text"
    if data_type == "n" or isinstance(value, (int, float)):
        return "number"

    return "general"


def format_value_with_context(value: Any, dna: CellDNA) -> str:
    """Format a cell value using its DNA for human-readable + LLM-friendly output.

    Instead of just "1234567.89", returns "$1,234,567.89 (USD)"
    Instead of just "0.15", returns "15.0% (percentage)"
    """
    if value is None:
        return ""

    st = dna.semantic_type

    if st.startswith("currency"):
        symbol = {"currency_usd": "$", "currency_eur": "€", "currency_gbp": "£",
                  "currency_jpy": "¥", "currency_inr": "₹"}.get(st, "$")
        try:
            num = float(value)
            formatted = f"{symbol}{num:,.2f}"
            if dna.formula:
                formatted += f" [={dna.formula}]"
            return formatted
        except (ValueError, TypeError):
            return str(value)

    if st == "accounting":
        try:
            num = float(value)
            return f"{num:,.2f}" if num >= 0 else f"({abs(num):,.2f})"
        except (ValueError, TypeError):
            return str(value)

    if st == "percentage":
        try:
            num = float(value)
            # Excel stores percentages as decimals (0.15 = 15%)
            if abs(num) <= 1.0:
                return f"{num * 100:.1f}%"
            return f"{num:.1f}%"
        except (ValueError, TypeError):
            return str(value)

    if st in ("date", "datetime"):
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")
        return str(value)

    if st == "time":
        if hasattr(value, "strftime"):
            return value.strftime("%H:%M:%S")
        return str(value)

    if st == "scientific":
        try:
            return f"{float(value):.2e}"
        except (ValueError, TypeError):
            return str(value)

    # Default: add formula context if present
    result = str(value)
    if dna.formula:
        result += f" [={dna.formula}]"
    if dna.comment:
        result += f" (note: {dna.comment})"
    return result


# ======================================================================
# Internal helpers
# ======================================================================

def _type_to_role(semantic_type: str) -> str | None:
    """Map a semantic type to a column role for the enricher.

    Returns None if the type is ambiguous (let the LLM decide).
    """
    role_map = {
        "currency_usd": "metric",
        "currency_eur": "metric",
        "currency_gbp": "metric",
        "currency_jpy": "metric",
        "currency_inr": "metric",
        "currency_krw": "metric",
        "currency": "metric",
        "accounting": "metric",
        "percentage": "metric",
        "decimal": "metric",
        "scientific": "metric",
        "integer": None,  # could be metric or identifier — ambiguous
        "date": "temporal",
        "time": "temporal",
        "datetime": "temporal",
        "text": None,  # could be identifier, descriptive, categorical
        "boolean": "categorical",
    }
    return role_map.get(semantic_type)


def _color_to_hex(color) -> str | None:
    """Extract 6-char hex color from an openpyxl Color object."""
    if color is None:
        return None
    if hasattr(color, "rgb") and color.rgb:
        rgb = str(color.rgb)
        # openpyxl stores as AARRGGBB (8 chars) or RRGGBB (6 chars)
        if len(rgb) == 8:
            hex6 = rgb[2:]  # strip alpha
        elif len(rgb) == 6:
            hex6 = rgb
        else:
            return None
        if hex6 in ("000000", "FFFFFF", "00000000"):
            return None  # treat black/white as "no color"
        return hex6
    return None


def _col_letter(col_index: int) -> str:
    """Convert 1-based column index to Excel letter (1=A, 26=Z, 27=AA)."""
    result = ""
    while col_index > 0:
        col_index, remainder = divmod(col_index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _coord_in_range(coord: str, cell_range: str) -> bool:
    """Check if a cell coordinate like 'B5' falls within a range like 'B2:B100'."""
    try:
        match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", cell_range)
        if not match:
            return False
        c_match = re.match(r"([A-Z]+)(\d+)", coord)
        if not c_match:
            return False

        start_col, start_row = match.group(1), int(match.group(2))
        end_col, end_row = match.group(3), int(match.group(4))
        cell_col, cell_row = c_match.group(1), int(c_match.group(2))

        return (start_col <= cell_col <= end_col and
                start_row <= cell_row <= end_row)
    except Exception:
        return False
