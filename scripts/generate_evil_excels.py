"""Generate 3 'Evil Spreadsheet' test files for parser stress-testing."""

import openpyxl
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "/Users/saiteja/Documents/Dev/EDMS/scripts/test_excels"

import os
os.makedirs(OUTPUT_DIR, exist_ok=True)


# ── Test 1: Floating Table (data doesn't start at A1) ──────────────
def create_floating_table():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Junk / logo area at the top
    ws["A1"] = "ACME Corp"
    ws["A2"] = "Confidential"
    ws["B3"] = "Generated: 2024-03-15"

    # Actual table starts at C5 (not A1!)
    ws["C5"] = "Report Title: Q3 Revenue Breakdown"

    # Headers at row 6, starting from column C
    headers = ["Employee ID", "Full Name", "Department", "Revenue ($)", "Region"]
    for i, h in enumerate(headers):
        ws.cell(row=6, column=3 + i, value=h)

    # Data rows
    data = [
        [101, "Alice Johnson", "Engineering", 78000, "West"],
        [102, "Bob Martinez", "Marketing", 45000, "East"],
        [103, "Carol Zhang", "Engineering", 92000, "West"],
        [104, "David Kim", "Sales", 61000, "Central"],
        [105, "Eve Wilson", "Marketing", None, "East"],  # missing revenue
    ]
    for r, row_data in enumerate(data):
        for c, val in enumerate(row_data):
            ws.cell(row=7 + r, column=3 + c, value=val)

    # More junk below the table
    ws["A15"] = "Footer: Do not distribute"

    path = f"{OUTPUT_DIR}/evil_1_floating_table.xlsx"
    wb.save(path)
    print(f"Created: {path}")


# ── Test 2: Merged Headers ─────────────────────────────────────────
def create_merged_headers():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Directory"

    # Row 1: top-level merged headers
    ws["A1"] = "ID"
    ws["B1"] = "Name"
    ws.merge_cells("C1:D1")
    ws["C1"] = "Contact Info"        # spans Phone + Email
    ws.merge_cells("E1:F1")
    ws["E1"] = "Address"             # spans City + State

    # Row 2: actual column headers
    ws["A2"] = "ID"
    ws["B2"] = "Name"
    ws["C2"] = "Phone"
    ws["D2"] = "Email"
    ws["E2"] = "City"
    ws["F2"] = "State"

    # Data
    data = [
        [1, "Alice", "555-0101", "alice@acme.com", "Portland", "OR"],
        [2, "Bob", "555-0102", "bob@acme.com", "Seattle", "WA"],
        [3, "Carol", "555-0103", None, "Austin", "TX"],  # missing email
        [4, "David", "555-0104", "david@acme.com", "Denver", "CO"],
    ]
    for r, row_data in enumerate(data):
        for c, val in enumerate(row_data):
            ws.cell(row=3 + r, column=1 + c, value=val)

    path = f"{OUTPUT_DIR}/evil_2_merged_headers.xlsx"
    wb.save(path)
    print(f"Created: {path}")


# ── Test 3: Hidden Sheet ───────────────────────────────────────────
def create_hidden_sheet():
    wb = openpyxl.Workbook()

    # Visible sheet: clean data
    ws1 = wb.active
    ws1.title = "Public Data"
    ws1["A1"] = "Product"
    ws1["B1"] = "Price"
    ws1["C1"] = "Stock"
    products = [
        ["Widget A", 29.99, 150],
        ["Widget B", 49.99, 80],
        ["Widget C", 19.99, 300],
    ]
    for r, row_data in enumerate(products):
        for c, val in enumerate(row_data):
            ws1.cell(row=2 + r, column=1 + c, value=val)

    # Hidden sheet: sensitive/junk data that should NOT be indexed
    ws2 = wb.create_sheet("Internal Costs")
    ws2.sheet_state = "hidden"
    ws2["A1"] = "Product"
    ws2["B1"] = "True Cost"
    ws2["C1"] = "Margin %"
    ws2["D1"] = "Supplier Secret Code"
    costs = [
        ["Widget A", 5.00, 0.83, "SUP-SECRET-001"],
        ["Widget B", 12.00, 0.76, "SUP-SECRET-002"],
        ["Widget C", 3.00, 0.85, "SUP-SECRET-003"],
    ]
    for r, row_data in enumerate(costs):
        for c, val in enumerate(row_data):
            ws2.cell(row=2 + r, column=1 + c, value=val)

    # Another visible sheet
    ws3 = wb.create_sheet("Categories")
    ws3["A1"] = "Category"
    ws3["B1"] = "Description"
    ws3["A2"] = "Electronics"
    ws3["B2"] = "Gadgets and devices"
    ws3["A3"] = "Home"
    ws3["B3"] = "Household items"

    path = f"{OUTPUT_DIR}/evil_3_hidden_sheet.xlsx"
    wb.save(path)
    print(f"Created: {path}")


if __name__ == "__main__":
    create_floating_table()
    create_merged_headers()
    create_hidden_sheet()
    print("\nAll 3 evil test files created!")
