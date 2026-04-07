"""Excel parser — 3-tier extraction: LlamaParse → Gemini → Spatial.

Tier 1: LlamaParse (best for clean tables, paid)
Tier 2: Gemini (best for complex/mixed layouts, free tier available)
Tier 3: Spatial algorithm + LLM fallback (free, offline)

All tiers output the same DocumentStructure interface used by the PDF
pipeline, so ChunkingPipeline, ContextEnricher, and SearchEngine work unchanged.
"""

from __future__ import annotations

import argparse
import json
import os
import re
from collections import deque
from dataclasses import asdict
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

from cell_dna import WorkbookDNA, SheetDNA, ColumnDNA, format_value_with_context
from pdf_processor import DocumentStructure, StructuredPage, StructuredSection


# ──────────────────────────────────────────────────────────────
# LlamaParse integration
# ──────────────────────────────────────────────────────────────

_LLAMA_PARSE_AVAILABLE = False
try:
    from llama_parse import LlamaParse as _LlamaParse
    _LLAMA_PARSE_AVAILABLE = True
except ImportError:
    pass

_GEMINI_AVAILABLE = False
try:
    from google import genai as _genai
    _GEMINI_AVAILABLE = True
except ImportError:
    pass


EXCEL_EXTENSIONS = {".xlsx", ".xls", ".csv"}


def is_excel_file(filename: str) -> bool:
    """Check by extension only. Use detect_file_type() for content-based detection."""
    return Path(filename).suffix.lower() in EXCEL_EXTENSIONS


# Magic bytes for file type detection
_MAGIC_SIGNATURES = {
    b"%PDF": "pdf",
    b"PK\x03\x04": "xlsx",  # ZIP-based (xlsx, docx, etc.)
    b"\xd0\xcf\x11\xe0": "xls",  # OLE2 (legacy .xls, .doc, .ppt)
}


def detect_file_type(file_path: str) -> str:
    """Detect actual file type from content magic bytes, ignoring the extension.

    Returns: "pdf", "xlsx", "xls", "csv", or "unknown"
    """
    path = Path(file_path)
    if not path.exists():
        return "unknown"

    # Read first 8 bytes for magic signature
    with open(path, "rb") as f:
        header = f.read(8)

    if not header:
        return "unknown"

    for magic, ftype in _MAGIC_SIGNATURES.items():
        if header[:len(magic)] == magic:
            # ZIP could be .xlsx or .docx — try to verify it's Excel
            if ftype == "xlsx":
                try:
                    import zipfile
                    with zipfile.ZipFile(path) as zf:
                        names = zf.namelist()
                        if any("xl/" in n or "xl\\" in n for n in names):
                            return "xlsx"
                        # Not an Excel ZIP (could be .docx, .pptx, etc.)
                        return "unknown"
                except Exception:
                    return "unknown"
            return ftype

    # No magic match — check if it looks like CSV (text with commas/tabs)
    ext = path.suffix.lower()
    if ext == ".csv":
        return "csv"
    try:
        sample = header + open(path, "rb").read(500)
        if sample.isascii() or sample.decode("utf-8", errors="ignore"):
            # Could be CSV if extension says so
            if ext in (".csv", ".tsv", ".txt"):
                return "csv"
    except Exception:
        pass

    return "unknown"


def smart_detect(file_path: str) -> str:
    """Detect whether file is 'excel' or 'pdf', regardless of extension.

    Returns: "excel", "pdf", or "unknown"
    """
    ftype = detect_file_type(file_path)
    if ftype in ("xlsx", "xls", "csv"):
        return "excel"
    if ftype == "pdf":
        return "pdf"
    # Fallback to extension
    ext = Path(file_path).suffix.lower()
    if ext in EXCEL_EXTENSIONS:
        return "excel"
    if ext == ".pdf":
        return "pdf"
    return "unknown"


# ──────────────────────────────────────────────────────────────
# Data classes for spatial regions
# ──────────────────────────────────────────────────────────────

@dataclass
class BoundingBox:
    r1: int  # row start (0-indexed)
    c1: int  # col start
    r2: int  # row end (inclusive)
    c2: int  # col end (inclusive)

    @property
    def width(self) -> int:
        return self.c2 - self.c1 + 1

    @property
    def height(self) -> int:
        return self.r2 - self.r1 + 1

    @property
    def area(self) -> int:
        return self.width * self.height


@dataclass
class DetectedRegion:
    bbox: BoundingBox
    region_type: str  # "table", "text", "heading", "key_value"
    cells: list  # list of (row, col, value) tuples
    density: float = 0.0
    metadata: dict = field(default_factory=dict)


@dataclass
class ExtractionCell:
    """Canonical Stage 1 cell record with provenance and display context."""
    sheet_name: str
    address: str
    row: int
    col: int
    raw_value: object
    display_value: str
    data_type: str
    number_format: str
    formula: str | None = None
    cached_value: object = None
    comment: str | None = None
    hyperlink: str | None = None
    bold: bool = False
    italic: bool = False
    fill_color: str | None = None
    indent: int = 0
    horizontal_alignment: str | None = None
    is_merged: bool = False
    merged_range: str | None = None
    hidden_row: bool = False
    hidden_col: bool = False
    inferred_role: str | None = None
    formula_references: list[str] = field(default_factory=list)
    row_label: str | None = None
    section_path: list[str] = field(default_factory=list)
    column_header_path: list[str] = field(default_factory=list)
    unit_context: str | None = None
    semantic_text: str | None = None
    provenance: dict[str, object] = field(default_factory=dict)


@dataclass
class ExtractionRegion:
    """Meaningful area on a sheet discovered during Stage 1."""
    region_id: str
    sheet_name: str
    region_type: str
    bbox: tuple[int, int, int, int]
    title: str | None = None
    header_rows: list[int] = field(default_factory=list)
    section_path: list[str] = field(default_factory=list)
    cells: list[ExtractionCell] = field(default_factory=list)
    metadata: dict = field(default_factory=dict)


@dataclass
class ExtractionTable:
    """Normalized table extracted from a workbook region."""
    table_id: str
    sheet_name: str
    region_id: str
    title: str | None
    headers: list[str]
    header_paths: list[list[str]]
    rows: list[dict[str, object]]
    row_provenance: list[dict[str, object]]
    formulas: list[dict[str, object]]
    dataframe: pd.DataFrame | None = None
    metadata: dict = field(default_factory=dict)


@dataclass
class WorkbookExtractionResult:
    """Stage 1 output from openpyxl-first workbook extraction."""
    file_path: str
    workbook_name: str
    sheet_order: list[str]
    hidden_sheets: list[str]
    named_ranges: list[dict[str, object]]
    sheet_metadata: dict[str, dict[str, object]] = field(default_factory=dict)
    regions: list[ExtractionRegion] = field(default_factory=list)
    tables: list[ExtractionTable] = field(default_factory=list)
    cells_by_sheet: dict[str, list[ExtractionCell]] = field(default_factory=dict)
    workbook_dna: dict[str, SheetDNA] = field(default_factory=dict)

    def to_parse_outputs(
        self,
    ) -> tuple[dict[str, pd.DataFrame], dict[str, list[str]], dict[str, SheetDNA]]:
        """Adapt Stage 1 output to the legacy parser contract."""
        dataframes: dict[str, pd.DataFrame] = {}
        formulas: dict[str, list[str]] = {}
        dna: dict[str, SheetDNA] = {}
        seen_names: dict[str, int] = {}

        for table in self.tables:
            if table.dataframe is None or table.dataframe.empty:
                continue

            clean_name = ExcelParser._clean_name(table.sheet_name)
            count = seen_names.get(clean_name, 0)
            seen_names[clean_name] = count + 1
            if count:
                clean_name = f"{clean_name}_{count}"

            df = table.dataframe.copy()
            dataframes[clean_name] = df
            formulas[clean_name] = [
                f"{f['cell']}: {f['formula']}"
                for f in table.formulas
                if f.get("cell") and f.get("formula")
            ]

            matched_dna = None
            if table.sheet_name in self.workbook_dna:
                matched_dna = self.workbook_dna[table.sheet_name]
            else:
                target = ExcelParser._clean_name(table.sheet_name)
                for sheet_name, sheet_dna in self.workbook_dna.items():
                    if ExcelParser._clean_name(sheet_name) == target:
                        matched_dna = sheet_dna
                        break
            if matched_dna:
                dna[clean_name] = matched_dna

        return dataframes, formulas, dna

    def to_dict(
        self,
        max_cells_per_sheet: int = 200,
        max_rows_per_table: int = 25,
    ) -> dict[str, object]:
        """Serialize the Stage 1 result to a JSON-friendly dict."""
        return {
            "file_path": self.file_path,
            "workbook_name": self.workbook_name,
            "sheet_order": self.sheet_order,
            "hidden_sheets": self.hidden_sheets,
            "named_ranges": self.named_ranges,
            "sheet_metadata": self.sheet_metadata,
            "regions": [
                {
                    **asdict(region),
                    "cells": [asdict(cell) for cell in region.cells[:max_cells_per_sheet]],
                    "cell_count": len(region.cells),
                }
                for region in self.regions
            ],
            "tables": [
                {
                    "table_id": table.table_id,
                    "sheet_name": table.sheet_name,
                    "region_id": table.region_id,
                    "title": table.title,
                    "headers": table.headers,
                    "header_paths": table.header_paths,
                    "rows": table.rows[:max_rows_per_table],
                    "row_count": len(table.rows),
                    "row_provenance": table.row_provenance[:max_rows_per_table],
                    "formula_count": len(table.formulas),
                    "formulas": table.formulas,
                    "metadata": table.metadata,
                }
                for table in self.tables
            ],
            "cells_by_sheet": {
                sheet_name: {
                    "cell_count": len(cells),
                    "cells": [asdict(cell) for cell in cells[:max_cells_per_sheet]],
                }
                for sheet_name, cells in self.cells_by_sheet.items()
            },
            "workbook_dna": {
                sheet_name: sheet_dna.to_dict()
                for sheet_name, sheet_dna in self.workbook_dna.items()
            },
        }


# ──────────────────────────────────────────────────────────────
# Main parser
# ──────────────────────────────────────────────────────────────

class ExcelParser:
    """Excel parser — LlamaParse primary, spatial fallback."""

    def __init__(self) -> None:
        self.last_stage1_result: WorkbookExtractionResult | None = None

    def extract_structure(self, file_path: str) -> DocumentStructure:
        """Main entry point. Tries LlamaParse first, falls back to spatial parser.

        Returns DocumentStructure (same interface as PDF pipeline).
        """
        path = Path(file_path)

        if path.suffix.lower() == ".csv":
            return self._parse_csv(file_path)

        # Tier 1: Gemini (best accuracy for complex layouts)
        google_key = os.environ.get("GOOGLE_API_KEY")
        if google_key and _GEMINI_AVAILABLE:
            try:
                result = self._parse_with_gemini(file_path, google_key)
                if result and result.pages:
                    print(f"[ExcelParser] Gemini succeeded: {len(result.pages)} pages")
                    return result
            except Exception as e:
                print(f"[ExcelParser] Gemini failed: {e}")

        # Fallback: Local spatial parser + LLM refinement
        print(f"[ExcelParser] Using spatial parser for {path.name}")

        # Read raw cell grid for all sheets (used by LLM refinement)
        raw_sheets = self._read_raw_cells(file_path)

        if path.suffix.lower() == ".xls":
            structure = self._parse_xls_fallback(file_path)
        else:
            structure = self._parse_spatial(file_path)

        # Phase 3: LLM refinement — send raw cell coordinates to LLM
        # for pages where the spatial parser struggled
        structure = self._llm_refine(structure, raw_sheets)
        return structure

    # ══════════════════════════════════════════════════════════
    # LlamaParse (primary)
    # ══════════════════════════════════════════════════════════

    def _parse_with_llamaparse(
        self, file_path: str, api_key: str
    ) -> DocumentStructure:
        """Parse Excel via LlamaParse cloud API → Markdown → DocumentStructure."""
        parser = _LlamaParse(
            api_key=api_key,
            result_type="markdown",
            verbose=False,
        )

        # LlamaParse returns a list of Document objects with .text (markdown)
        documents = parser.load_data(file_path)

        if not documents:
            return DocumentStructure(
                title=Path(file_path).stem,
                doc_type="spreadsheet",
                pages=[],
                font_profile={},
                source_profile={"mode": "llamaparse", "parser": "llamaparse"},
            )

        # Convert LlamaParse markdown output to DocumentStructure
        pages = []
        for page_num, doc in enumerate(documents, start=1):
            sections = self._markdown_to_sections(doc.text)
            if sections:
                pages.append(StructuredPage(
                    page_number=page_num,
                    sections=sections,
                ))

        return DocumentStructure(
            title=Path(file_path).stem,
            doc_type="spreadsheet",
            pages=pages,
            font_profile={},
            source_profile={"mode": "llamaparse", "parser": "llamaparse"},
        )

    # ══════════════════════════════════════════════════════════
    # Gemini (Tier 2)
    # ══════════════════════════════════════════════════════════

    def _parse_with_gemini(
        self, file_path: str, api_key: str
    ) -> DocumentStructure:
        """Parse Excel via Gemini — send cell grid with coordinates.

        Gemini doesn't accept Excel file uploads, so we read the raw cell
        grid (A1: value, B3: value) and send it as text. Gemini's large
        context window and strong reasoning extract structure from the
        spatial coordinates.
        """
        # Read raw cell grids per sheet
        raw_sheets = self._read_raw_cells(file_path)
        if not raw_sheets:
            return DocumentStructure(
                title=Path(file_path).stem, doc_type="spreadsheet",
                pages=[], font_profile={},
                source_profile={"mode": "gemini", "parser": "gemini-2.5-flash"},
            )

        client = _genai.Client(api_key=api_key)
        pages = []

        import time
        for page_num, cell_grid in raw_sheets.items():
            if not cell_grid.strip():
                continue

            # Rate limit: avoid bursting Gemini free tier
            if page_num > 1:
                time.sleep(2)

            # Truncate to 8000 chars per sheet (Gemini has huge context but keep it focused)
            grid_text = cell_grid[:8000]

            prompt = (
                "You are a document extraction engine.\n\n"
                "Below is a cell grid from one sheet of an Excel spreadsheet. "
                "Each line shows the cell coordinate and value "
                "(e.g., 'B12: FALBERG KATHRYN E' = column B, row 12).\n\n"
                "The SPATIAL POSITION is critical:\n"
                "- Cells in the same ROW but different columns are often label → value pairs\n"
                "- Cells in the same COLUMN share a category\n"
                "- A cell like 'B10: 1. Name and Address' is a LABEL for the value in B12\n\n"
                "TASK: Extract ALL information as clean Markdown:\n"
                "- Use ## for section headings\n"
                "- Use | col1 | col2 | for data tables (flatten multi-row headers into ONE header row)\n"
                "- Use **Key:** Value for key-value pairs\n"
                "- Use plain text for paragraphs\n\n"
                "RULES:\n"
                "- Extract EVERY value. Do not skip or summarize.\n"
                "- Preserve exact values: $30.45, 102828, 3/5/2026\n"
                "- Pair labels with their values using spatial proximity\n"
                "- Do NOT include cell coordinates in the output\n"
                "- Do NOT explain what you're doing — just output the extracted content\n\n"
                f"CELL GRID:\n{grid_text}"
            )

            try:
                response = client.models.generate_content(
                    model="gemini-2.5-flash",
                    contents=prompt,
                )
                markdown = response.text or ""
            except Exception as e:
                print(f"[ExcelParser] Gemini failed on page {page_num}: {e}")
                continue

            if markdown.strip():
                sections = self._markdown_to_sections(markdown)
                if sections:
                    pages.append(StructuredPage(
                        page_number=page_num,
                        sections=sections,
                    ))

        return DocumentStructure(
            title=Path(file_path).stem,
            doc_type="spreadsheet",
            pages=pages,
            font_profile={},
            source_profile={"mode": "gemini", "parser": "gemini-2.5-flash"},
        )

    def _markdown_to_sections(self, markdown: str) -> list[StructuredSection]:
        """Convert LlamaParse markdown output to StructuredSections.

        LlamaParse returns markdown with:
        - # headings
        - | table | rows |
        - Plain text paragraphs
        """
        sections = []
        current_heading = ""
        current_block: list[str] = []
        current_type = "text"  # "text" or "table"

        def flush_block():
            nonlocal current_block, current_type
            if not current_block:
                return
            content = "\n".join(current_block)
            if content.strip():
                # If it's a table, also try to extract a DataFrame
                lines = [{"text": line, "bbox": None}
                         for line in current_block if line.strip()]

                if current_type == "table":
                    try:
                        df = self._markdown_table_to_df(content)
                        if df is not None:
                            lines.append({"__meta__": {
                                "dataframe": df,
                                "sheet_name": current_heading or "sheet",
                                "formulas": [],
                                "schema": self._generate_schema(
                                    df, current_heading or "sheet"
                                ),
                            }})
                    except Exception:
                        pass  # DataFrame extraction is best-effort

                sections.append(StructuredSection(
                    title=current_heading or "content",
                    level=3,
                    content=content,
                    lines=lines,
                    section_type=current_type,
                ))
            current_block = []
            current_type = "text"

        for line in markdown.split("\n"):
            stripped = line.strip()

            # Heading
            heading_match = re.match(r"^(#{1,3})\s+(.+)$", stripped)
            if heading_match:
                flush_block()
                level = len(heading_match.group(1))
                current_heading = heading_match.group(2).strip()
                sections.append(StructuredSection(
                    title=current_heading,
                    level=level,
                    content=current_heading,
                    lines=[],
                    section_type="text",
                ))
                continue

            # Table row
            if stripped.startswith("|") and stripped.endswith("|"):
                if current_type != "table":
                    flush_block()
                    current_type = "table"
                current_block.append(stripped)
                continue

            # Separator line (|---|---|)
            if re.match(r"^\|[\s\-:|]+\|$", stripped):
                if current_type == "table":
                    current_block.append(stripped)
                continue

            # Empty line — flush
            if not stripped:
                flush_block()
                continue

            # Regular text
            if current_type != "text":
                flush_block()
                current_type = "text"
            current_block.append(stripped)

        flush_block()
        return sections

    def _markdown_table_to_df(self, markdown: str) -> pd.DataFrame | None:
        """Convert a markdown table string to a pandas DataFrame."""
        try:
            lines = [l.strip() for l in markdown.split("\n")
                     if l.strip() and l.strip().startswith("|")]
            if len(lines) < 2:
                return None

            # Parse header
            header = [c.strip() for c in lines[0].strip("|").split("|")]

            # Skip separator line
            data_lines = [l for l in lines[1:]
                          if not re.match(r"^\|[\s\-:|]+\|$", l)]

            rows = []
            for line in data_lines:
                cells = [c.strip() for c in line.strip("|").split("|")]
                # Pad or trim to match header width
                if len(cells) < len(header):
                    cells.extend([""] * (len(header) - len(cells)))
                elif len(cells) > len(header):
                    cells = cells[:len(header)]
                rows.append(cells)

            if not rows:
                return None

            df = pd.DataFrame(rows, columns=header)

            # Clean column names
            clean_cols = [
                re.sub(r"[^a-z0-9]+", "_", str(c).strip().lower()).strip("_") or f"col_{i}"
                for i, c in enumerate(df.columns)
            ]
            seen: dict[str, int] = {}
            deduped = []
            for col in clean_cols:
                if col in seen:
                    seen[col] += 1
                    deduped.append(f"{col}_{seen[col]}")
                else:
                    seen[col] = 0
                    deduped.append(col)
            df.columns = deduped

            # Infer numeric types
            for col in df.columns:
                try:
                    df[col] = pd.to_numeric(df[col])
                except (ValueError, TypeError):
                    pass

            return df if not df.empty else None
        except Exception:
            return None  # DataFrame extraction is best-effort

    # ══════════════════════════════════════════════════════════
    # Spatial parser (fallback)
    # ══════════════════════════════════════════════════════════

    def _parse_spatial(self, file_path: str) -> DocumentStructure:
        """Local spatial parser — grid binarization → island detection → classification."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        wb_formulas = None
        try:
            wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        except Exception:
            pass

        pages = []
        for page_num, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            if ws.sheet_state != "visible":
                continue

            # Phase 1: Spatial boundary detection
            grid, cell_values = self._build_grid(ws)
            if not grid:
                continue

            islands = self._extract_islands(grid)
            if not islands:
                continue

            # Phase 2: Classify each island
            regions = []
            for bbox in islands:
                cells = self._extract_cells(cell_values, bbox)
                region = self._classify_region(bbox, cells, ws)
                regions.append(region)

            # Sort regions top-to-bottom, left-to-right
            regions.sort(key=lambda r: (r.bbox.r1, r.bbox.c1))

            # Convert regions to StructuredSections
            sections = self._regions_to_sections(
                regions, sheet_name, ws, wb_formulas
            )

            if sections:
                pages.append(StructuredPage(
                    page_number=page_num,
                    sections=sections,
                ))

        wb.close()
        if wb_formulas:
            wb_formulas.close()

        return DocumentStructure(
            title=Path(file_path).stem,
            doc_type="spreadsheet",
            pages=pages,
            font_profile={},
            source_profile={"mode": "excel_spatial", "parser": "openpyxl"},
        )

    def parse(self, file_path: str) -> tuple[dict[str, pd.DataFrame], dict[str, list[str]], dict[str, SheetDNA]]:
        """Parse Excel file into DataFrames + formulas + CellDNA metadata.

        Returns:
            dataframes: {sheet_name: DataFrame}
            formulas:   {sheet_name: [formula_strings]}
            dna:        {sheet_name: SheetDNA} — the hidden truth from openpyxl
        """
        path = Path(file_path)

        if path.suffix.lower() == ".xlsx":
            stage1 = self.extract_stage1(file_path)
            self.last_stage1_result = stage1
            dataframes, formulas, dna = stage1.to_parse_outputs()
            if dataframes:
                return dataframes, formulas, dna

        structure = self.extract_structure(file_path)
        dataframes = {}
        formulas = {}

        for page in structure.pages:
            for section in page.sections:
                if section.section_type == "table":
                    meta = section.metadata if hasattr(section, "metadata") else {}
                    # Try to get the metadata dict from lines (we store it there)
                    section_meta = {}
                    if section.lines and isinstance(section.lines[-1], dict):
                        section_meta = section.lines[-1].get("__meta__", {})

                    df = section_meta.get("dataframe")
                    sheet_name = section_meta.get("sheet_name", section.title)
                    table_formulas = section_meta.get("formulas", [])

                    if df is not None:
                        clean_name = self._clean_name(sheet_name)
                        # Deduplicate names
                        base = clean_name
                        i = 1
                        while clean_name in dataframes:
                            clean_name = f"{base}_{i}"
                            i += 1
                        dataframes[clean_name] = df
                        formulas[clean_name] = table_formulas

        # If no tables found, fall back to reading as flat DataFrame
        if not dataframes:
            dfs, fms = self._fallback_parse(file_path)
            dna = self._extract_dna_safe(file_path) if dfs else {}
            return dfs, fms, dna

        # Clean HTML entities (e.g. &nbsp;) from all string columns
        for name, df in dataframes.items():
            for col in df.select_dtypes(include=["object"]).columns:
                df[col] = df[col].apply(
                    lambda v: re.sub(r"&nbsp;", " ", str(v)).strip()
                    if isinstance(v, str) else v
                )
            dataframes[name] = df

        # Extract CellDNA for all sheets
        dna = self._extract_dna_safe(file_path)

        return dataframes, formulas, dna

    def extract_stage1(self, file_path: str) -> WorkbookExtractionResult:
        """Run the openpyxl-first Stage 1 extraction for modern Excel files."""
        path = Path(file_path)
        if path.suffix.lower() != ".xlsx":
            return self._fallback_stage1(file_path)

        wb_values = openpyxl.load_workbook(file_path, data_only=True)
        wb_formulas = None
        try:
            wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        except Exception:
            wb_formulas = None

        workbook_dna = self._extract_dna_safe(file_path)
        hidden_sheets = [
            ws.title for ws in wb_values.worksheets
            if ws.sheet_state != "visible"
        ]
        result = WorkbookExtractionResult(
            file_path=file_path,
            workbook_name=path.name,
            sheet_order=[ws.title for ws in wb_values.worksheets],
            hidden_sheets=hidden_sheets,
            named_ranges=self._extract_named_ranges(wb_values),
            workbook_dna=workbook_dna,
        )

        try:
            for sheet_index, ws in enumerate(wb_values.worksheets, start=1):
                ws_formula = wb_formulas[ws.title] if wb_formulas and ws.title in wb_formulas.sheetnames else None
                merge_map = self._build_merge_map(ws)
                cell_lookup = self._extract_sheet_cells(ws, ws_formula, merge_map)
                self._enrich_sheet_cell_context(ws, cell_lookup, merge_map)
                cells = list(cell_lookup.values())
                result.cells_by_sheet[ws.title] = cells
                result.sheet_metadata[ws.title] = {
                    "sheet_index": sheet_index,
                    "sheet_state": ws.sheet_state,
                    "used_range": self._sheet_used_range(ws),
                    "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
                    "frozen_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
                    "hidden_rows": [
                        idx for idx, dim in ws.row_dimensions.items() if dim.hidden
                    ],
                    "hidden_cols": [
                        key for key, dim in ws.column_dimensions.items() if dim.hidden
                    ],
                }

                if ws.sheet_state != "visible":
                    continue

                table_bboxes: list[BoundingBox] = []
                for table_idx, table in enumerate(ws.tables.values(), start=1):
                    bbox = self._bbox_from_range(table.ref)
                    table_bboxes.append(bbox)
                    region_id = f"{self._clean_name(ws.title)}:table:{table_idx}"
                    region_cells = self._cells_in_bbox(cell_lookup, bbox)
                    title = getattr(table, "displayName", None) or table.name
                    region = ExtractionRegion(
                        region_id=region_id,
                        sheet_name=ws.title,
                        region_type="table",
                        bbox=(bbox.r1 + 1, bbox.c1 + 1, bbox.r2 + 1, bbox.c2 + 1),
                        title=title,
                        header_rows=[bbox.r1 + 1],
                        section_path=[title] if title else [],
                        cells=region_cells,
                        metadata={"source": "excel_table", "ref": table.ref},
                    )
                    result.regions.append(region)
                    extracted = self._build_extraction_table(
                        ws, ws_formula, bbox, region_id, title, region_cells, source="excel_table"
                    )
                    if extracted:
                        result.tables.append(extracted)

                grid, cell_values = self._build_grid(ws)
                if not grid:
                    continue

                islands = self._extract_islands(grid)
                if not islands:
                    continue

                current_heading = ws.title
                heuristic_index = 0
                for bbox in sorted(islands, key=lambda b: (b.r1, b.c1)):
                    if any(self._bbox_overlaps(bbox, tb) for tb in table_bboxes):
                        continue
                    detected_cells = self._extract_cells(cell_values, bbox)
                    detected = self._classify_region(bbox, detected_cells, ws)
                    heuristic_index += 1
                    region_id = f"{self._clean_name(ws.title)}:{detected.region_type}:{heuristic_index}"
                    region_cells = self._cells_in_bbox(cell_lookup, bbox)
                    title = current_heading
                    if detected.region_type == "heading":
                        title = self._heading_text(detected)
                        current_heading = title

                    region = ExtractionRegion(
                        region_id=region_id,
                        sheet_name=ws.title,
                        region_type=detected.region_type,
                        bbox=(bbox.r1 + 1, bbox.c1 + 1, bbox.r2 + 1, bbox.c2 + 1),
                        title=title,
                        header_rows=[bbox.r1 + 1] if detected.region_type == "table" else [],
                        section_path=[current_heading] if current_heading and current_heading != ws.title else [],
                        cells=region_cells,
                        metadata={"source": "heuristic", "density": detected.density},
                    )
                    result.regions.append(region)

                    if detected.region_type == "table":
                        extracted = self._build_extraction_table(
                            ws, ws_formula, bbox, region_id, title, region_cells, source="heuristic"
                        )
                        if extracted:
                            result.tables.append(extracted)
        finally:
            wb_values.close()
            if wb_formulas:
                wb_formulas.close()

        return result

    def _extract_dna_safe(self, file_path: str) -> dict[str, SheetDNA]:
        """Extract WorkbookDNA, returning empty dict on failure.

        CellDNA extraction is best-effort — the pipeline works without it,
        just with less metadata for the enricher.
        """
        ext = Path(file_path).suffix.lower()
        if ext in (".csv", ".xls"):
            # CSV has no styles, .xls needs xlrd (no openpyxl support)
            return {}
        try:
            with WorkbookDNA(file_path) as wdna:
                return dict(wdna.sheets)
        except Exception as e:
            print(f"[ExcelParser] CellDNA extraction failed (non-fatal): {e}")
            return {}

    def _fallback_stage1(self, file_path: str) -> WorkbookExtractionResult:
        """Best-effort Stage 1 output for non-xlsx files."""
        dataframes, formulas = self._fallback_parse(file_path)
        workbook_dna = self._extract_dna_safe(file_path) if dataframes else {}
        result = WorkbookExtractionResult(
            file_path=file_path,
            workbook_name=Path(file_path).name,
            sheet_order=list(dataframes.keys()),
            hidden_sheets=[],
            named_ranges=[],
            workbook_dna=workbook_dna,
        )
        for idx, (sheet_name, df) in enumerate(dataframes.items(), start=1):
            table_id = f"{self._clean_name(sheet_name)}:table:1"
            result.sheet_metadata[sheet_name] = {
                "sheet_index": idx,
                "sheet_state": "visible",
                "used_range": None,
                "merged_ranges": [],
                "frozen_panes": None,
                "hidden_rows": [],
                "hidden_cols": [],
            }
            table = ExtractionTable(
                table_id=table_id,
                sheet_name=sheet_name,
                region_id=table_id,
                title=sheet_name,
                headers=list(df.columns),
                header_paths=[[sheet_name, c] for c in df.columns],
                rows=df.to_dict(orient="records"),
                row_provenance=[
                    {"row_index": row_idx, "source_cells": {}, "section_path": [sheet_name]}
                    for row_idx in range(len(df))
                ],
                formulas=[
                    {"cell": item.split(": ", 1)[0], "formula": item.split(": ", 1)[1], "references": []}
                    for item in formulas.get(sheet_name, [])
                    if ": " in item
                ],
                dataframe=df.copy(),
                metadata={"source": "fallback"},
            )
            result.tables.append(table)
        return result

    def _extract_named_ranges(self, wb) -> list[dict[str, object]]:
        """Return workbook-level defined names in a serializable form."""
        named_ranges: list[dict[str, object]] = []
        try:
            for name, defined_name in wb.defined_names.items():
                destinations = []
                try:
                    for sheet_name, coord in defined_name.destinations:
                        destinations.append({"sheet": sheet_name, "coord": coord})
                except Exception:
                    pass
                named_ranges.append(
                    {
                        "name": name,
                        "value": getattr(defined_name, "value", None),
                        "destinations": destinations,
                    }
                )
        except Exception:
            pass
        return named_ranges

    def _build_merge_map(self, ws) -> dict[str, str]:
        """Map each coordinate in merged ranges to its parent range string."""
        merge_map: dict[str, str] = {}
        for merge_range in ws.merged_cells.ranges:
            for row in ws[str(merge_range)]:
                for cell in row:
                    merge_map[cell.coordinate] = str(merge_range)
        return merge_map

    def _extract_sheet_cells(self, ws, ws_formula, merge_map: dict[str, str]) -> dict[tuple[int, int], ExtractionCell]:
        """Extract all populated or semantically meaningful cells from a sheet."""
        cells: dict[tuple[int, int], ExtractionCell] = {}
        for row in ws.iter_rows():
            for cell in row:
                formula_value = ws_formula[cell.coordinate].value if ws_formula else None
                comment = cell.comment.text if cell.comment else None
                hyperlink = cell.hyperlink.target if cell.hyperlink else None
                has_signal = any(
                    [
                        cell.value is not None and str(cell.value).strip(),
                        isinstance(formula_value, str) and formula_value.startswith("="),
                        comment,
                        hyperlink,
                        cell.has_style,
                    ]
                )
                if not has_signal:
                    continue

                fill_color = None
                try:
                    fill_color = cell.fill.fgColor.rgb
                except Exception:
                    fill_color = None

                extraction = ExtractionCell(
                    sheet_name=ws.title,
                    address=cell.coordinate,
                    row=cell.row,
                    col=cell.column,
                    raw_value=cell.value,
                    display_value="" if cell.value is None else str(cell.value),
                    data_type=cell.data_type,
                    number_format=cell.number_format or "General",
                    formula=formula_value if isinstance(formula_value, str) and formula_value.startswith("=") else None,
                    cached_value=cell.value,
                    comment=comment,
                    hyperlink=hyperlink,
                    bold=bool(cell.font and cell.font.bold),
                    italic=bool(cell.font and cell.font.italic),
                    fill_color=fill_color,
                    indent=int(getattr(cell.alignment, "indent", 0) or 0),
                    horizontal_alignment=getattr(cell.alignment, "horizontal", None),
                    is_merged=cell.coordinate in merge_map,
                    merged_range=merge_map.get(cell.coordinate),
                    hidden_row=bool(ws.row_dimensions[cell.row].hidden),
                    hidden_col=bool(ws.column_dimensions[get_column_letter(cell.column)].hidden),
                    inferred_role=self._infer_cell_role(cell.value, formula_value, cell.number_format),
                    formula_references=self._extract_formula_references(formula_value),
                )
                cells[(cell.row - 1, cell.column - 1)] = extraction
        return cells

    def _build_extraction_table(
        self,
        ws,
        ws_formula,
        bbox: BoundingBox,
        region_id: str,
        title: str | None,
        region_cells: list[ExtractionCell],
        source: str,
    ) -> ExtractionTable | None:
        """Convert a detected table region into the Stage 1 normalized table model."""
        data = []
        for r in range(bbox.r1, bbox.r2 + 1):
            row_data = []
            for c in range(bbox.c1, bbox.c2 + 1):
                row_data.append(ws.cell(row=r + 1, column=c + 1).value)
            data.append(row_data)

        if not data:
            return None

        raw_df = pd.DataFrame(data)
        if raw_df.empty:
            return None

        header_row = 0
        for idx in range(min(3, len(raw_df))):
            row = raw_df.iloc[idx]
            str_count = sum(1 for value in row if isinstance(value, str) and value and value.strip())
            if str_count >= len(row) * 0.5:
                header_row = idx
                break

        raw_headers = raw_df.iloc[header_row].tolist()
        clean_headers = []
        seen_headers: dict[str, int] = {}
        for col_idx, header in enumerate(raw_headers):
            clean = re.sub(r"[^a-z0-9]+", "_", str(header).strip().lower()).strip("_") or f"col_{col_idx}"
            count = seen_headers.get(clean, 0)
            seen_headers[clean] = count + 1
            if count:
                clean = f"{clean}_{count}"
            clean_headers.append(clean)

        body_df = raw_df.iloc[header_row + 1:].reset_index(drop=True)
        keep_rows = body_df.notna().any(axis=1)
        keep_cols = []
        keep_headers = []
        keep_col_positions = []
        for idx, header in enumerate(clean_headers):
            series = body_df.iloc[:, idx] if idx < len(body_df.columns) else pd.Series(dtype=object)
            if series.notna().any() and not re.match(r"^(nan|none)(_\d+)?$", header):
                keep_cols.append(idx)
                keep_headers.append(header)
                keep_col_positions.append(idx)

        if not keep_cols:
            return None

        kept_row_positions = [idx for idx, flag in enumerate(keep_rows.tolist()) if flag]
        df = body_df.loc[keep_rows, keep_cols].copy()
        if df.empty:
            return None

        df.columns = keep_headers
        df = df.reset_index(drop=True)

        for col in df.columns:
            try:
                if df[col].dtype == "object":
                    df[col] = pd.to_numeric(df[col])
            except (ValueError, TypeError):
                pass

        row_provenance = []
        formula_entries = []
        for output_row_index, body_row_index in enumerate(kept_row_positions):
            excel_row = bbox.r1 + header_row + 2 + body_row_index
            source_cells = {}
            for header_name, rel_col_idx in zip(keep_headers, keep_col_positions):
                excel_col = bbox.c1 + rel_col_idx + 1
                coordinate = f"{get_column_letter(excel_col)}{excel_row}"
                source_cells[header_name] = coordinate
                if ws_formula:
                    formula_cell = ws_formula[coordinate]
                    if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                        formula_entries.append(
                            {
                                "cell": coordinate,
                                "column": header_name,
                                "row_index": output_row_index,
                                "formula": formula_cell.value,
                                "references": self._extract_formula_references(formula_cell.value),
                            }
                        )
            row_provenance.append(
                {
                    "row_index": output_row_index,
                    "excel_row": excel_row,
                    "source_cells": source_cells,
                    "section_path": [title] if title else [ws.title],
                }
            )

        return ExtractionTable(
            table_id=region_id,
            sheet_name=ws.title,
            region_id=region_id,
            title=title,
            headers=keep_headers,
            header_paths=[[title or ws.title, header] for header in keep_headers],
            rows=df.to_dict(orient="records"),
            row_provenance=row_provenance,
            formulas=formula_entries,
            dataframe=df,
            metadata={
                "source": source,
                "bbox": (bbox.r1 + 1, bbox.c1 + 1, bbox.r2 + 1, bbox.c2 + 1),
                "header_row": bbox.r1 + header_row + 1,
                "cell_count": len(region_cells),
            },
        )

    def _enrich_sheet_cell_context(
        self,
        ws,
        cell_lookup: dict[tuple[int, int], ExtractionCell],
        merge_map: dict[str, str],
    ) -> None:
        """Attach row/column/section meaning to extracted cells."""
        unit_context = self._detect_unit_context(ws, merge_map)
        section_rows = self._detect_section_rows(ws, merge_map)

        for (row_idx, col_idx), cell in cell_lookup.items():
            if cell.hidden_row or cell.hidden_col:
                continue

            row_label, row_label_coord = self._resolve_row_label(ws, row_idx + 1, col_idx + 1, merge_map)
            column_header_path, header_coords = self._resolve_column_header_path(
                ws, row_idx + 1, col_idx + 1, merge_map
            )
            section_path = self._resolve_section_path(
                ws, row_idx + 1, col_idx + 1, merge_map, section_rows
            )

            cell.row_label = row_label
            cell.column_header_path = column_header_path
            cell.section_path = section_path
            cell.unit_context = unit_context
            cell.provenance = {
                "row_header_cell": row_label_coord,
                "column_header_cells": header_coords,
                "section_header_cells": [
                    item["coord"] for item in section_rows
                    if item["value"] in section_path
                ],
            }
            cell.semantic_text = self._build_semantic_text(cell)

    def _cells_in_bbox(
        self,
        cell_lookup: dict[tuple[int, int], ExtractionCell],
        bbox: BoundingBox,
    ) -> list[ExtractionCell]:
        """Return extracted cells within a bounding box."""
        return [
            cell
            for (row, col), cell in cell_lookup.items()
            if bbox.r1 <= row <= bbox.r2 and bbox.c1 <= col <= bbox.c2
        ]

    def _bbox_from_range(self, ref: str) -> BoundingBox:
        """Convert an Excel range like A1:D10 into a 0-based bounding box."""
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        return BoundingBox(
            r1=min_row - 1,
            c1=min_col - 1,
            r2=max_row - 1,
            c2=max_col - 1,
        )

    def _bbox_overlaps(self, left: BoundingBox, right: BoundingBox) -> bool:
        """Return True when two bounding boxes overlap."""
        return not (
            left.r2 < right.r1
            or right.r2 < left.r1
            or left.c2 < right.c1
            or right.c2 < left.c1
        )

    def _sheet_used_range(self, ws) -> str | None:
        """Return the used range in A1 notation."""
        if not ws.max_row or not ws.max_column:
            return None
        return f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    def _infer_cell_role(
        self,
        value: object,
        formula_value: object,
        number_format: str | None,
    ) -> str | None:
        """Lightweight structural role inference for Stage 1 cells."""
        text = str(value).strip().lower() if value is not None else ""
        fmt = (number_format or "").lower()
        if isinstance(formula_value, str) and formula_value.startswith("="):
            if any(keyword in text for keyword in ("total", "subtotal", "net")):
                return "total"
            return "computed"
        if any(token in fmt for token in ("yy", "dd", "mm", "h:", "m/")):
            return "temporal"
        if "%" in fmt:
            return "percentage"
        if any(symbol in fmt for symbol in ("$", "€", "£", "¥", "₹")):
            return "currency"
        if isinstance(value, str) and text.endswith(":"):
            return "key_value_label"
        if isinstance(value, str):
            return "label"
        if isinstance(value, (int, float)):
            return "metric"
        return None

    def _extract_formula_references(self, formula_value: object) -> list[str]:
        """Return direct cell/range references from a formula string."""
        if not isinstance(formula_value, str) or not formula_value.startswith("="):
            return []
        refs = re.findall(
            r"(?:'[^']+'!|[A-Za-z0-9_]+!)?\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?",
            formula_value,
        )
        seen = set()
        ordered = []
        for ref in refs:
            if ref not in seen:
                seen.add(ref)
                ordered.append(ref)
        return ordered

    def _effective_cell_value(self, ws, row: int, col: int, merge_map: dict[str, str]) -> object:
        """Return the visible value for a coordinate, resolving merged parents."""
        coordinate = f"{get_column_letter(col)}{row}"
        cell = ws.cell(row=row, column=col)
        value = cell.value
        if value not in (None, ""):
            return value

        merged_range = merge_map.get(coordinate)
        if not merged_range:
            return value

        min_col, min_row, _, _ = range_boundaries(merged_range)
        return ws.cell(row=min_row, column=min_col).value

    def _effective_cell_coordinate(self, row: int, col: int, merge_map: dict[str, str]) -> str:
        """Return the anchor coordinate for merged cells, else the direct coordinate."""
        coordinate = f"{get_column_letter(col)}{row}"
        merged_range = merge_map.get(coordinate)
        if not merged_range:
            return coordinate
        min_col, min_row, _, _ = range_boundaries(merged_range)
        return f"{get_column_letter(min_col)}{min_row}"

    def _cell_has_numeric_right(self, ws, row: int, col: int, max_cols: int, merge_map: dict[str, str]) -> bool:
        """Heuristic: whether a row has numeric cells to the right of a label."""
        for candidate_col in range(col + 1, min(max_cols, col + 8) + 1):
            value = self._effective_cell_value(ws, row, candidate_col, merge_map)
            if isinstance(value, (int, float)):
                return True
        return False

    def _resolve_row_label(
        self,
        ws,
        row: int,
        col: int,
        merge_map: dict[str, str],
    ) -> tuple[str | None, str | None]:
        """Find the most likely row label to the left of a cell."""
        best_value = None
        best_coord = None
        for candidate_col in range(col - 1, 0, -1):
            value = self._effective_cell_value(ws, row, candidate_col, merge_map)
            if isinstance(value, str) and value.strip():
                text = value.strip()
                if text.startswith("(") and text.endswith(")"):
                    continue
                best_value = text
                best_coord = self._effective_cell_coordinate(row, candidate_col, merge_map)
                break
        return best_value, best_coord

    def _resolve_column_header_path(
        self,
        ws,
        row: int,
        col: int,
        merge_map: dict[str, str],
    ) -> tuple[list[str], list[str]]:
        """Collect top-down header values above the cell's column."""
        values: list[str] = []
        coords: list[str] = []
        seen: set[str] = set()
        for candidate_row in range(1, row):
            value = self._effective_cell_value(ws, candidate_row, col, merge_map)
            if not isinstance(value, str) or not value.strip():
                continue
            text = " ".join(value.strip().split())
            if len(text) > 120:
                continue
            lowered = text.lower()
            if lowered in seen:
                continue
            seen.add(lowered)
            values.append(text)
            coords.append(self._effective_cell_coordinate(candidate_row, col, merge_map))
        return values, coords

    def _detect_section_rows(
        self,
        ws,
        merge_map: dict[str, str],
    ) -> list[dict[str, object]]:
        """Detect section headers such as Operations / Financing / Investing."""
        sections: list[dict[str, object]] = []
        max_cols = min(ws.max_column or 0, 8)
        for row in range(1, (ws.max_row or 0) + 1):
            text_candidates: list[tuple[int, str, str]] = []
            numeric_count = 0
            for col in range(1, max_cols + 1):
                value = self._effective_cell_value(ws, row, col, merge_map)
                if isinstance(value, str) and value.strip():
                    coord = self._effective_cell_coordinate(row, col, merge_map)
                    if any(item[2] == coord for item in text_candidates):
                        continue
                    text_candidates.append((col, value.strip(), coord))
                elif isinstance(value, (int, float)):
                    numeric_count += 1

            if numeric_count > 0 or len(text_candidates) != 1:
                continue

            col, text, coord = text_candidates[0]
            cell = ws[coord]
            if (
                len(text) <= 60
                and self._cell_has_numeric_right(ws, row + 1, col, ws.max_column or col, merge_map)
                and (bool(cell.font and cell.font.bold) or text.istitle())
            ):
                sections.append({"row": row, "value": text, "coord": coord})
        return sections

    def _resolve_section_path(
        self,
        ws,
        row: int,
        col: int,
        merge_map: dict[str, str],
        section_rows: list[dict[str, object]],
    ) -> list[str]:
        """Return the nearest structural section above the current row."""
        matches = [item for item in section_rows if item["row"] < row]
        if not matches:
            return []
        return [matches[-1]["value"]]

    def _detect_unit_context(self, ws, merge_map: dict[str, str]) -> str | None:
        """Find top-of-sheet unit text such as '(In millions)'."""
        max_rows = min(ws.max_row or 0, 8)
        max_cols = min(ws.max_column or 0, 6)
        for row in range(1, max_rows + 1):
            for col in range(1, max_cols + 1):
                value = self._effective_cell_value(ws, row, col, merge_map)
                if isinstance(value, str):
                    text = value.strip()
                    if text.startswith("(") and text.endswith(")"):
                        return text
        return None

    def _build_semantic_text(self, cell: ExtractionCell) -> str | None:
        """Build a human-readable meaning string for the extracted cell."""
        parts = []
        if cell.sheet_name:
            parts.append(f"Sheet: {cell.sheet_name}")
        if cell.section_path:
            parts.append(f"Section: {' > '.join(cell.section_path)}")
        if cell.row_label:
            parts.append(f"Line item: {cell.row_label}")
        if cell.column_header_path:
            parts.append(f"Column path: {' > '.join(cell.column_header_path)}")
        if cell.unit_context:
            parts.append(f"Units: {cell.unit_context}")
        if cell.display_value:
            parts.append(f"Value: {cell.display_value}")
        return " | ".join(parts) if parts else None

    # ══════════════════════════════════════════════════════════
    # PHASE 1: Spatial Boundary Detection
    # ══════════════════════════════════════════════════════════

    def _build_grid(self, ws) -> tuple[list[list[int]], dict]:
        """Convert worksheet to binary grid + cell value map.

        Returns:
            grid: 2D list of 0s and 1s
            cell_values: dict mapping (row, col) → cell value
        """
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            return [], {}

        grid = [[0] * max_col for _ in range(max_row)]
        cell_values = {}

        for row in ws.iter_rows(min_row=1, max_row=max_row,
                                min_col=1, max_col=max_col):
            for cell in row:
                r = cell.row - 1  # 0-indexed
                c = cell.column - 1

                has_data = False
                if cell.value is not None and str(cell.value).strip():
                    has_data = True

                # Also check for significant formatting
                if not has_data and cell.has_style:
                    font = cell.font
                    if font and font.bold:
                        has_data = bool(cell.value is not None)

                if has_data:
                    grid[r][c] = 1
                    cell_values[(r, c)] = cell.value

        return grid, cell_values

    def _extract_islands(self, grid: list[list[int]]) -> list[BoundingBox]:
        """Find contiguous islands of 1s using BFS flood fill.

        Two cells are contiguous if adjacent horizontally, vertically, or diagonally.
        Returns bounding boxes for each island.
        """
        if not grid:
            return []

        rows = len(grid)
        cols = len(grid[0])
        visited = [[False] * cols for _ in range(rows)]
        islands = []

        for r in range(rows):
            for c in range(cols):
                if grid[r][c] == 1 and not visited[r][c]:
                    # BFS to find all connected cells
                    bbox = self._bfs_island(grid, visited, r, c, rows, cols)
                    if bbox.area >= 1:  # at least 1 cell
                        islands.append(bbox)

        return islands

    def _bfs_island(
        self, grid, visited, start_r, start_c, rows, cols
    ) -> BoundingBox:
        """BFS flood fill from a starting cell. Returns bounding box of the island."""
        queue = deque([(start_r, start_c)])
        visited[start_r][start_c] = True
        min_r, max_r = start_r, start_r
        min_c, max_c = start_c, start_c

        # 8-directional adjacency (horizontal, vertical, diagonal)
        directions = [
            (-1, -1), (-1, 0), (-1, 1),
            (0, -1),           (0, 1),
            (1, -1),  (1, 0),  (1, 1),
        ]

        while queue:
            r, c = queue.popleft()
            min_r = min(min_r, r)
            max_r = max(max_r, r)
            min_c = min(min_c, c)
            max_c = max(max_c, c)

            for dr, dc in directions:
                nr, nc = r + dr, c + dc
                if 0 <= nr < rows and 0 <= nc < cols:
                    if grid[nr][nc] == 1 and not visited[nr][nc]:
                        visited[nr][nc] = True
                        queue.append((nr, nc))

        return BoundingBox(r1=min_r, c1=min_c, r2=max_r, c2=max_c)

    def _extract_cells(
        self, cell_values: dict, bbox: BoundingBox
    ) -> list[tuple[int, int, object]]:
        """Get all non-empty cells within a bounding box."""
        cells = []
        for r in range(bbox.r1, bbox.r2 + 1):
            for c in range(bbox.c1, bbox.c2 + 1):
                if (r, c) in cell_values:
                    cells.append((r, c, cell_values[(r, c)]))
        return cells

    # ══════════════════════════════════════════════════════════
    # PHASE 2: Heuristic Classification
    # ══════════════════════════════════════════════════════════

    def _classify_region(
        self, bbox: BoundingBox, cells: list, ws
    ) -> DetectedRegion:
        """Classify a detected island as table, text, heading, or key_value."""
        if not cells:
            return DetectedRegion(bbox=bbox, region_type="text", cells=cells)

        # Calculate density
        density = len(cells) / max(bbox.area, 1)

        # Gather stats
        values = [v for _, _, v in cells]
        str_values = [v for v in values if isinstance(v, str)]
        num_values = [v for v in values if isinstance(v, (int, float))]
        avg_str_len = (
            sum(len(s) for s in str_values) / len(str_values)
            if str_values else 0
        )

        # ── Heading: 1 row, 1-2 cells, short text ──
        if bbox.height == 1 and len(cells) <= 2 and avg_str_len < 80:
            # Check for bold or uppercase
            is_bold = False
            is_upper = False
            for r, c, v in cells:
                cell_obj = ws.cell(row=r + 1, column=c + 1)
                if cell_obj.font and cell_obj.font.bold:
                    is_bold = True
                if isinstance(v, str) and v.strip().isupper() and len(v.strip()) > 2:
                    is_upper = True
            if is_bold or is_upper or (bbox.width <= 2 and len(cells) == 1):
                return DetectedRegion(
                    bbox=bbox, region_type="heading",
                    cells=cells, density=density,
                )

        # ── Key-Value Pairs: exactly 2 columns, col1 often ends with ":" ──
        if bbox.width == 2 and bbox.height >= 2:
            col1_vals = [v for r, c, v in cells if c == bbox.c1]
            colon_count = sum(
                1 for v in col1_vals
                if isinstance(v, str) and v.strip().endswith(":")
            )
            if colon_count >= len(col1_vals) * 0.4 and len(col1_vals) >= 2:
                return DetectedRegion(
                    bbox=bbox, region_type="key_value",
                    cells=cells, density=density,
                )

        # ── Data Table: 3+ columns, density > 0.4, type consistency ──
        if bbox.width >= 3 and density > 0.4:
            # Check column-wise type consistency
            col_types = {}
            for r, c, v in cells:
                col_types.setdefault(c, []).append(type(v).__name__)

            # If at least one column is mostly numeric and one is mostly string
            has_numeric_col = False
            has_string_col = False
            for col_vals in col_types.values():
                numeric_ratio = sum(1 for t in col_vals if t in ("int", "float")) / max(len(col_vals), 1)
                string_ratio = sum(1 for t in col_vals if t == "str") / max(len(col_vals), 1)
                if numeric_ratio > 0.5:
                    has_numeric_col = True
                if string_ratio > 0.5:
                    has_string_col = True

            if has_numeric_col or (bbox.height >= 3 and density > 0.5):
                return DetectedRegion(
                    bbox=bbox, region_type="table",
                    cells=cells, density=density,
                )

        # ── Also classify as table if wide enough with good density ──
        if bbox.width >= 3 and bbox.height >= 3 and density > 0.3:
            return DetectedRegion(
                bbox=bbox, region_type="table",
                cells=cells, density=density,
            )

        # ── Text: 1-2 columns, or long strings ──
        if avg_str_len > 40 or (bbox.width <= 2 and len(str_values) > len(num_values)):
            return DetectedRegion(
                bbox=bbox, region_type="text",
                cells=cells, density=density,
            )

        # ── Default: text for small regions, table for larger ones ──
        if bbox.area <= 6:
            return DetectedRegion(
                bbox=bbox, region_type="text",
                cells=cells, density=density,
            )

        return DetectedRegion(
            bbox=bbox, region_type="table",
            cells=cells, density=density,
        )

    # ══════════════════════════════════════════════════════════
    # Convert regions to StructuredSections
    # ══════════════════════════════════════════════════════════

    def _regions_to_sections(
        self,
        regions: list[DetectedRegion],
        sheet_name: str,
        ws,
        wb_formulas,
    ) -> list[StructuredSection]:
        """Convert detected regions into StructuredSection objects."""
        sections = []
        current_heading = sheet_name

        for region in regions:
            if region.region_type == "heading":
                current_heading = self._heading_text(region)
                sections.append(StructuredSection(
                    title=current_heading,
                    level=1,
                    content=current_heading,
                    lines=[],
                    section_type="text",
                ))

            elif region.region_type == "text":
                text = self._text_content(region)
                if text.strip():
                    sections.append(StructuredSection(
                        title=current_heading,
                        level=3,
                        content=text,
                        lines=[{"text": line, "bbox": None} for line in text.split("\n") if line.strip()],
                        section_type="text",
                    ))

            elif region.region_type == "key_value":
                text = self._kvp_content(region)
                if text.strip():
                    sections.append(StructuredSection(
                        title=current_heading,
                        level=3,
                        content=text,
                        lines=[{"text": line, "bbox": None} for line in text.split("\n") if line.strip()],
                        section_type="text",
                    ))

            elif region.region_type == "table":
                md_content, df, table_formulas = self._table_content(
                    region, ws, wb_formulas, sheet_name
                )
                if md_content.strip():
                    # Store DataFrame in lines metadata for backward compat
                    lines = [{"text": line, "bbox": None} for line in md_content.split("\n") if line.strip()]
                    lines.append({"__meta__": {
                        "dataframe": df,
                        "sheet_name": sheet_name,
                        "formulas": table_formulas,
                        "schema": self._generate_schema(df, sheet_name) if df is not None else "",
                    }})

                    sections.append(StructuredSection(
                        title=current_heading,
                        level=3,
                        content=md_content,
                        lines=lines,
                        section_type="table",
                    ))

        return sections

    # ══════════════════════════════════════════════════════════
    # Content extraction per region type
    # ══════════════════════════════════════════════════════════

    def _heading_text(self, region: DetectedRegion) -> str:
        values = [str(v).strip() for _, _, v in region.cells if v is not None]
        return " ".join(values)

    def _text_content(self, region: DetectedRegion) -> str:
        """Concatenate text cells into clean paragraphs."""
        rows: dict[int, list] = {}
        for r, c, v in region.cells:
            rows.setdefault(r, []).append((c, v))

        lines = []
        for r in sorted(rows.keys()):
            row_cells = sorted(rows[r], key=lambda x: x[0])
            text = " ".join(str(v).strip() for _, v in row_cells if v is not None)
            if text.strip():
                lines.append(text.strip())

        return "\n".join(lines)

    def _kvp_content(self, region: DetectedRegion) -> str:
        """Convert key-value pairs to semantic strings."""
        rows: dict[int, list] = {}
        for r, c, v in region.cells:
            rows.setdefault(r, []).append((c, v))

        lines = []
        for r in sorted(rows.keys()):
            row_cells = sorted(rows[r], key=lambda x: x[0])
            if len(row_cells) >= 2:
                key = str(row_cells[0][1]).strip().rstrip(":")
                val = str(row_cells[1][1]).strip()
                lines.append(f"{key}: {val}")
            elif len(row_cells) == 1:
                lines.append(str(row_cells[0][1]).strip())

        return "\n".join(lines)

    def _table_content(
        self, region: DetectedRegion, ws, wb_formulas, sheet_name: str
    ) -> tuple[str, pd.DataFrame | None, list[str]]:
        """Extract table as Markdown + DataFrame + formulas."""
        bbox = region.bbox

        # Read the sub-region into a DataFrame
        data = []
        for r in range(bbox.r1, bbox.r2 + 1):
            row_data = []
            for c in range(bbox.c1, bbox.c2 + 1):
                cell = ws.cell(row=r + 1, column=c + 1)
                row_data.append(cell.value)
            data.append(row_data)

        if not data:
            return "", None, []

        df = pd.DataFrame(data)

        # Detect header row (first row that's mostly strings)
        header_row = 0
        for i in range(min(3, len(df))):
            row = df.iloc[i]
            str_count = sum(1 for v in row if isinstance(v, str) and v and v.strip())
            if str_count >= len(row) * 0.5:
                header_row = i
                break

        # Set header
        if len(df) > header_row:
            headers = df.iloc[header_row].astype(str).tolist()
            df = df.iloc[header_row + 1:].reset_index(drop=True)
            df.columns = headers

        # Clean column names
        clean_cols = [
            re.sub(r"[^a-z0-9]+", "_", str(c).strip().lower()).strip("_") or f"col_{i}"
            for i, c in enumerate(df.columns)
        ]
        # Deduplicate
        seen: dict[str, int] = {}
        deduped = []
        for col in clean_cols:
            if col in seen:
                seen[col] += 1
                deduped.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                deduped.append(col)
        df.columns = deduped

        # Drop empty rows/cols
        df = df.dropna(how="all").dropna(axis=1, how="all")

        # Drop nan-named columns
        nan_cols = [c for c in df.columns if re.match(r"^(nan|none)(_\d+)?$", c)]
        if nan_cols:
            df = df.drop(columns=nan_cols, errors="ignore")

        if df.empty:
            return "", None, []

        # Infer types
        for col in df.columns:
            try:
                if df[col].dtype == "object":
                    df[col] = pd.to_numeric(df[col])
            except (ValueError, TypeError):
                pass

        df = df.reset_index(drop=True)

        # Extract formulas for this region
        table_formulas = []
        if wb_formulas:
            try:
                ws_f = wb_formulas[ws.title]
                for r in range(bbox.r1, bbox.r2 + 1):
                    for c in range(bbox.c1, bbox.c2 + 1):
                        cell = ws_f.cell(row=r + 1, column=c + 1)
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            table_formulas.append(f"{cell.coordinate}: {cell.value}")
            except Exception:
                pass

        # Generate Markdown table
        md_lines = []
        if not df.empty:
            # Header row
            md_lines.append("| " + " | ".join(str(c) for c in df.columns) + " |")
            md_lines.append("| " + " | ".join("---" for _ in df.columns) + " |")
            # Data rows
            for _, row in df.iterrows():
                vals = []
                for c in df.columns:
                    v = row[c]
                    vals.append(str(v) if pd.notna(v) else "")
                md_lines.append("| " + " | ".join(vals) + " |")

        return "\n".join(md_lines), df if not df.empty else None, table_formulas

    def _generate_schema(self, df: pd.DataFrame, table_name: str) -> str:
        """Auto-generate schema description for LLM SQL prompts."""
        if df is None or df.empty:
            return ""

        lines = [f"Table: {table_name}", f"Rows: {len(df)}"]
        for col in df.columns:
            dtype = str(df[col].dtype)
            non_null = int(df[col].notna().sum())
            unique = df[col].dropna().unique()
            num_unique = len(unique)
            sample = unique[:5].tolist()
            sample = [str(v) if hasattr(v, "isoformat") else v for v in sample]

            desc = f"  - {col} ({dtype}, {non_null}/{len(df)} non-null, {num_unique} unique)"
            if num_unique <= 10:
                desc += f" — values: {sample}"
            else:
                desc += f" — e.g.: {sample}"
            if df[col].dtype in ("int64", "float64", "int32", "float32"):
                desc += f" — range: [{df[col].min()}, {df[col].max()}]"
            lines.append(desc)

        return "\n".join(lines)

    # ══════════════════════════════════════════════════════════
    # ══════════════════════════════════════════════════════════
    # PHASE 3: LLM Refinement for ambiguous regions
    # ══════════════════════════════════════════════════════════

    def _read_raw_cells(self, file_path: str) -> dict[int, str]:
        """Read every non-empty cell with its coordinate from all sheets.

        Returns {page_number: "A1: value\\nB3: value\\n..."} — the raw cell
        grid that preserves spatial layout for LLM refinement.
        """
        raw_sheets: dict[int, str] = {}
        ext = Path(file_path).suffix.lower()

        try:
            if ext == ".xls":
                sheets = pd.read_excel(file_path, sheet_name=None, header=None, engine="xlrd")
                for page_num, (name, df) in enumerate(sheets.items(), start=1):
                    lines = []
                    for r_idx, row in df.iterrows():
                        for c_idx, val in enumerate(row):
                            if pd.notna(val) and str(val).strip():
                                col_letter = self._index_to_excel_col(c_idx)
                                lines.append(f"{col_letter}{r_idx + 1}: {val}")
                    raw_sheets[page_num] = "\n".join(lines)
            else:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                page_num = 0
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if ws.sheet_state != "visible":
                        continue
                    page_num += 1
                    lines = []
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value is not None and str(cell.value).strip():
                                lines.append(f"{cell.coordinate}: {cell.value}")
                    raw_sheets[page_num] = "\n".join(lines)
                wb.close()
        except Exception:
            pass

        return raw_sheets

    @staticmethod
    def _index_to_excel_col(idx: int) -> str:
        """Convert 0-based column index to Excel letter (0=A, 25=Z, 26=AA)."""
        result = ""
        idx += 1
        while idx > 0:
            idx, remainder = divmod(idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _llm_refine(
        self, structure: DocumentStructure, raw_sheets: dict[int, str] | None = None
    ) -> DocumentStructure:
        """Post-process: send raw cell grid to LLM for pages where spatial parser struggled.

        The key insight: sending cell coordinates (A1: value, B3: value) preserves
        the 2D spatial layout that gets lost when concatenating rows left-to-right.
        """
        groq_key = os.environ.get("GROQ_API_KEY")
        has_ollama = self._check_ollama()
        if not groq_key and not has_ollama:
            return structure

        if not raw_sheets:
            return structure

        refined_pages = []
        for page in structure.pages:
            if self._page_needs_refinement(page):
                cell_grid = raw_sheets.get(page.page_number, "")
                if cell_grid:
                    refined = self._refine_page_with_llm(
                        page, groq_key, has_ollama, cell_grid
                    )
                    if refined:
                        refined_pages.append(refined)
                        continue
            refined_pages.append(page)

        structure.pages = refined_pages
        return structure

    def _page_needs_refinement(self, page: StructuredPage) -> bool:
        """Check if a page's sections look fragmented or ambiguous."""
        sections = page.sections
        if not sections:
            return False

        # Too many heading-level sections (over-fragmented)
        headings = sum(1 for s in sections if s.level == 1)
        if headings > 6:
            return True

        # Tables with only 1 column (likely failed to detect structure)
        for s in sections:
            if s.section_type == "table" and s.content.count("|") < 10:
                return True

        # Very short text sections that look like table fragments
        short_texts = sum(
            1 for s in sections
            if s.section_type == "text" and s.level == 3
            and len(s.content) < 30
        )
        if short_texts > 5:
            return True

        return False

    def _check_ollama(self) -> bool:
        """Check if Ollama is running locally."""
        import httpx
        try:
            with httpx.Client(timeout=3) as client:
                resp = client.get("http://localhost:11434/api/tags")
                return resp.status_code == 200
        except Exception:
            return False

    def _refine_page_with_llm(
        self, page: StructuredPage, groq_key: str | None,
        has_ollama: bool, cell_grid: str = ""
    ) -> StructuredPage | None:
        """Send raw cell grid (with coordinates) to LLM for restructuring.

        The cell grid preserves spatial layout:
          A1: Created by EDGAR Online, Inc.
          B12: FALBERG KATHRYN E
          D12: Trade Desk, Inc. [ TTD ]
          F14: Director

        The LLM can see WHERE each value lives and reconstruct meaning.
        """
        import httpx

        if not cell_grid or len(cell_grid.strip()) < 20:
            return None

        # Truncate to fit context window
        cell_grid = cell_grid[:6000]

        prompt = (
            "You are a document structure extraction engine.\n\n"
            "Below is a cell grid from an Excel spreadsheet. Each line shows the "
            "cell coordinate and its value (e.g., 'B12: FALBERG KATHRYN E' means "
            "column B, row 12 contains that name).\n\n"
            "The spatial position matters — cells in the same row are related, "
            "and cells in the same column share a category.\n\n"
            "TASK: Analyze the spatial layout and extract ALL information into "
            "clean, structured sections.\n\n"
            "For each section, use this exact format:\n"
            "---SECTION---\n"
            "TYPE: one of [heading, table, text, key_value]\n"
            "TITLE: descriptive title\n"
            "CONTENT:\n"
            "For tables: use | col1 | col2 | markdown format\n"
            "For key_value: use Key: Value per line\n"
            "For text: clean paragraphs\n"
            "---END---\n\n"
            "IMPORTANT:\n"
            "- Extract EVERY piece of data — names, dates, amounts, addresses\n"
            "- For key-value pairs, pair the label with its value using spatial proximity\n"
            "- Cells in the same row but different columns are often label-value pairs\n"
            "- Preserve exact values — do not summarize or paraphrase\n\n"
            f"CELL GRID:\n{cell_grid}"
        )

        llm_output = None

        # Try Groq first
        if groq_key:
            try:
                with httpx.Client(timeout=30) as client:
                    resp = client.post(
                        "https://api.groq.com/openai/v1/chat/completions",
                        headers={
                            "Authorization": f"Bearer {groq_key}",
                            "Content-Type": "application/json",
                        },
                        json={
                            "model": "llama-3.3-70b-versatile",
                            "messages": [{"role": "user", "content": prompt}],
                            "max_tokens": 3000,
                            "temperature": 0,
                        },
                    )
                    if resp.status_code == 200:
                        llm_output = resp.json()["choices"][0]["message"]["content"]
                        print(f"[ExcelParser] Groq refined page {page.page_number}")
            except Exception:
                pass

        # Fallback to Ollama
        if not llm_output and has_ollama:
            try:
                with httpx.Client(timeout=60) as client:
                    resp = client.post(
                        "http://localhost:11434/api/generate",
                        json={
                            "model": "llama3.2",
                            "prompt": prompt,
                            "stream": False,
                        },
                    )
                    if resp.status_code == 200:
                        llm_output = resp.json().get("response")
                        print(f"[ExcelParser] Ollama refined page {page.page_number}")
            except Exception as e:
                print(f"[ExcelParser] LLM refinement failed page {page.page_number}: {e}")

        if not llm_output:
            return None

        # Parse LLM output into StructuredSections
        sections = self._parse_llm_sections(llm_output)
        if not sections:
            return None

        return StructuredPage(page_number=page.page_number, sections=sections)

    def _parse_llm_sections(self, llm_output: str) -> list[StructuredSection]:
        """Parse the LLM's structured output into StructuredSections."""
        sections = []
        blocks = re.split(r"---SECTION---", llm_output)

        for block in blocks:
            block = block.strip()
            if not block or block == "---END---":
                continue

            # Remove trailing ---END---
            block = re.sub(r"---END---.*$", "", block, flags=re.DOTALL).strip()

            # Extract TYPE, TITLE, CONTENT
            type_match = re.search(r"TYPE:\s*(\w+)", block)
            title_match = re.search(r"TITLE:\s*(.+)", block)
            content_match = re.search(r"CONTENT:\s*\n?(.*)", block, re.DOTALL)

            if not content_match:
                continue

            region_type = type_match.group(1).lower() if type_match else "text"
            title = title_match.group(1).strip() if title_match else "content"
            content = content_match.group(1).strip()

            if not content:
                continue

            # Map type to section_type and level
            section_type = "text"
            level = 3
            if region_type == "heading":
                level = 1
            elif region_type == "table":
                section_type = "table"
            elif region_type == "key_value":
                section_type = "text"  # KVPs stored as text

            lines = [{"text": line, "bbox": None}
                     for line in content.split("\n") if line.strip()]

            # If it's a table, try to extract DataFrame
            if section_type == "table":
                try:
                    df = self._markdown_table_to_df(content)
                    if df is not None:
                        lines.append({"__meta__": {
                            "dataframe": df,
                            "sheet_name": title,
                            "formulas": [],
                            "schema": self._generate_schema(df, title),
                        }})
                except Exception:
                    pass

            sections.append(StructuredSection(
                title=title,
                level=level,
                content=content,
                lines=lines,
                section_type=section_type,
            ))

        return sections

    # CSV + XLS + fallback
    # ══════════════════════════════════════════════════════════

    def _parse_xls_fallback(self, file_path: str) -> DocumentStructure:
        """Parse old .xls files using xlrd via pandas (openpyxl doesn't support .xls)."""
        sheets = pd.read_excel(file_path, sheet_name=None, header=None, engine="xlrd")
        pages = []

        for page_num, (sheet_name, df) in enumerate(sheets.items(), start=1):
            # Drop fully empty rows/cols
            df = df.dropna(how="all").dropna(axis=1, how="all")
            if df.empty:
                continue

            # Drop >80% null columns (padding)
            col_null = df.isna().sum() / len(df)
            df = df.loc[:, col_null <= 0.8]
            df = df.dropna(how="all").reset_index(drop=True)
            if df.empty:
                continue

            # Strip leading title rows (1 non-null cell)
            if len(df.columns) >= 3:
                while len(df) > 0:
                    if df.iloc[0].notna().sum() <= 1:
                        df = df.iloc[1:].reset_index(drop=True)
                    else:
                        break

            if df.empty:
                continue

            # Convert each row to text, classify blocks
            sections = []
            current_block: list[str] = []
            current_type = "text"

            for _, row in df.iterrows():
                non_null = row.dropna()
                if len(non_null) == 0:
                    continue
                line = " ".join(str(v).strip() for v in non_null if str(v).strip())
                if not line:
                    continue

                # Detect if this looks like a heading (single short value)
                if len(non_null) == 1 and len(line) < 60:
                    # Flush current block
                    if current_block:
                        content = "\n".join(current_block)
                        sections.append(StructuredSection(
                            title=sections[-1].title if sections else sheet_name,
                            level=3, content=content,
                            lines=[{"text": l, "bbox": None} for l in current_block],
                            section_type=current_type,
                        ))
                        current_block = []

                    sections.append(StructuredSection(
                        title=line, level=1, content=line,
                        lines=[], section_type="text",
                    ))
                else:
                    # Check if tabular (has numbers)
                    has_nums = any(isinstance(v, (int, float)) for v in non_null)
                    if has_nums:
                        current_type = "table"
                    current_block.append(line)

            # Flush remaining
            if current_block:
                content = "\n".join(current_block)
                sections.append(StructuredSection(
                    title=sections[-1].title if sections else sheet_name,
                    level=3, content=content,
                    lines=[{"text": l, "bbox": None} for l in current_block],
                    section_type=current_type,
                ))

            if sections:
                pages.append(StructuredPage(page_number=page_num, sections=sections))

        return DocumentStructure(
            title=Path(file_path).stem,
            doc_type="spreadsheet",
            pages=pages,
            font_profile={},
            source_profile={"mode": "excel_xls_fallback", "parser": "xlrd"},
        )

    def _parse_csv(self, file_path: str) -> DocumentStructure:
        """CSV files are always single-table documents."""
        df = pd.read_csv(file_path)
        df.columns = [
            re.sub(r"[^a-z0-9]+", "_", str(c).strip().lower()).strip("_") or f"col_{i}"
            for i, c in enumerate(df.columns)
        ]

        md_lines = ["| " + " | ".join(df.columns) + " |"]
        md_lines.append("| " + " | ".join("---" for _ in df.columns) + " |")
        for _, row in df.head(100).iterrows():
            md_lines.append("| " + " | ".join(str(v) if pd.notna(v) else "" for v in row) + " |")

        content = "\n".join(md_lines)
        lines = [{"text": line, "bbox": None} for line in md_lines]
        lines.append({"__meta__": {
            "dataframe": df,
            "sheet_name": Path(file_path).stem,
            "formulas": [],
            "schema": self._generate_schema(df, Path(file_path).stem),
        }})

        section = StructuredSection(
            title=Path(file_path).stem,
            level=3,
            content=content,
            lines=lines,
            section_type="table",
        )

        return DocumentStructure(
            title=Path(file_path).stem,
            doc_type="spreadsheet",
            pages=[StructuredPage(page_number=1, sections=[section])],
            font_profile={},
            source_profile={"mode": "csv", "parser": "pandas"},
        )

    def _fallback_parse(self, file_path: str) -> tuple[dict, dict]:
        """Fallback: read sheets as flat DataFrames (legacy behavior)."""
        try:
            sheets = pd.read_excel(file_path, sheet_name=None, engine="openpyxl")
            dfs = {}
            for name, df in sheets.items():
                clean = self._clean_name(name)
                df = df.dropna(how="all").dropna(axis=1, how="all")
                if not df.empty:
                    df.columns = [
                        re.sub(r"[^a-z0-9]+", "_", str(c).strip().lower()).strip("_") or f"col_{i}"
                        for i, c in enumerate(df.columns)
                    ]
                    dfs[clean] = df
            return dfs, {k: [] for k in dfs}
        except Exception:
            return {}, {}

    @staticmethod
    def _clean_name(name: str) -> str:
        cleaned = re.sub(r"[^a-z0-9]+", "_", str(name).strip().lower()).strip("_")
        return cleaned or "sheet"


def _build_cli_parser() -> argparse.ArgumentParser:
    """Return the command-line parser for inspecting Excel Stage 1 output."""
    parser = argparse.ArgumentParser(
        description="Inspect Stage 1 openpyxl extraction output for an Excel file.",
    )
    parser.add_argument("file_path", help="Path to .xlsx/.xls/.csv file")
    parser.add_argument(
        "--format",
        choices=("json", "summary"),
        default="summary",
        help="Output format",
    )
    parser.add_argument(
        "--max-cells",
        type=int,
        default=50,
        help="Maximum cells to include per sheet/region in JSON output",
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        default=20,
        help="Maximum table rows to include in JSON output",
    )
    return parser


def _format_stage1_summary(stage1: WorkbookExtractionResult) -> str:
    """Render a concise human-readable summary of the Stage 1 extraction."""
    lines = [
        f"Workbook: {stage1.workbook_name}",
        f"Sheets: {', '.join(stage1.sheet_order) if stage1.sheet_order else '(none)'}",
    ]
    if stage1.hidden_sheets:
        lines.append(f"Hidden sheets: {', '.join(stage1.hidden_sheets)}")
    if stage1.named_ranges:
        lines.append(f"Named ranges: {len(stage1.named_ranges)}")

    for sheet_name in stage1.sheet_order:
        meta = stage1.sheet_metadata.get(sheet_name, {})
        lines.append("")
        lines.append(f"[Sheet] {sheet_name}")
        lines.append(f"  State: {meta.get('sheet_state', 'unknown')}")
        lines.append(f"  Used range: {meta.get('used_range') or '(unknown)'}")
        lines.append(f"  Cell records: {len(stage1.cells_by_sheet.get(sheet_name, []))}")

        sheet_tables = [t for t in stage1.tables if t.sheet_name == sheet_name]
        lines.append(f"  Tables: {len(sheet_tables)}")
        for table in sheet_tables:
            lines.append(
                f"    - {table.table_id}: {len(table.rows)} rows, headers={table.headers}"
            )
            if table.formulas:
                sample_formula = table.formulas[0]
                lines.append(
                    f"      sample formula: {sample_formula.get('cell')} = {sample_formula.get('formula')}"
                )

        regions = [r for r in stage1.regions if r.sheet_name == sheet_name]
        if regions:
            region_counts: dict[str, int] = {}
            for region in regions:
                region_counts[region.region_type] = region_counts.get(region.region_type, 0) + 1
            counts = ", ".join(f"{kind}={count}" for kind, count in sorted(region_counts.items()))
            lines.append(f"  Regions: {counts}")

    return "\n".join(lines)


def main() -> int:
    """CLI entry point for inspecting Stage 1 extraction output."""
    parser = _build_cli_parser()
    args = parser.parse_args()

    excel_parser = ExcelParser()
    stage1 = excel_parser.extract_stage1(args.file_path)

    if args.format == "json":
        payload = stage1.to_dict(
            max_cells_per_sheet=max(1, args.max_cells),
            max_rows_per_table=max(1, args.max_rows),
        )
        print(json.dumps(payload, indent=2, default=str))
    else:
        print(_format_stage1_summary(stage1))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
